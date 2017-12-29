# -*- coding: utf-8 -*-
import traceback,json,sys,datetime,types,random,time,urllib2,urllib,cookielib,requests
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import get_template
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render,render_to_response
from django.template import loader,Context,RequestContext
from Business.models import *
from TaskAndFlow.models import *
from UserAndPrj.models import *
from django.db.models import Q
from django.core import serializers
from openpyxl import Workbook
from openpyxl import load_workbook
from Scc4PM.settings import UPLOAD_DIR
from dss.Serializer import serializer
from django.core.paginator import Paginator
from PIL import Image
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from Scc4PM.settings import CURRENT_PROJECT_ID,DATABASES
from django.db.models import *
from decimal import Decimal
from TaskAndFlow.utility_filemanager import *
from Business.serializers import *
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.authentication import SessionAuthentication
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import renderer_classes
from rest_framework.renderers import StaticHTMLRenderer
import rest_framework_filters
'''execl处理工具'''

'''数据格式处理方法'''
def dataformat(data,datatype='float'):
    result = None
    if datatype==None:'float'
    if datatype=='float':
        if data==None:data=0
        result = float(data)
    return result
'''预算execl管理'''
def execlhandle(docid,pactid,operate):
    response_data = {}
    docobj = Document.objects.get(id = docid)
    dicitem = docobj.docdirectory.all()[0]
    path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
    wb = load_workbook(path)
    sheetname = wb.get_sheet_names()[0]
    sheetContent = wb[sheetname]
    response_data['sheetname'] = sheetname
    row = sheetContent.rows
    listqd = []
    qingdan = {}
    currlistqd = []
    queryList = []
    curr_time = datetime.datetime.now()
    curr_time = curr_time.strftime('%Y-%m-%d %H:%M:%S')

    if operate==1:
        for index,cell in enumerate(row):
            if index>4:
                if cell[0].value != None:
                    if index != 5:
                        qingdan['dinges'] = currlistqd
                        listqd.append(qingdan)
                        currlistqd = []
                        qingdan = {}
                    currqd = {}
                    qingdan['xuhao'] = cell[0].value
                    qingdan['code'] = cell[1].value
                    qingdan['name'] = cell[2].value
                    qingdan['unit'] = cell[3].value
                    designBqs = str(cell[4].value).strip()
                    if designBqs==None or designBqs=='None':
                        designBqs = '0'
                    buildBqs = '0'
                    qingdan['nums'] = float(designBqs)
                    qingdan['jsnums'] = float(buildBqs)
                    price1 = str(cell[5].value).strip()
                    if price1==None or price1=='None':
                        price1 = '0'
                    price2 = str(cell[7].value).strip()
                    if price2==None or price2=='None':
                        price2 = '0'
                    qingdan['price1'] = Decimal(float(price1)).quantize(Decimal('0.00'))
                    qingdan['price2'] = Decimal(float(price2)).quantize(Decimal('0.00'))
                    qingdan['jsprice'] = 0
                    if BqItem.objects.filter(BQItem_Code=str((qingdan['code']).strip())).count()>0:
                        currbqitem = BqItem.objects.filter(BQItem_Code=str((qingdan['code']).strip()))[0]
                    else:
                        currbqitem = BqItem.objects.create(BQItem_Code=str((qingdan['code']).strip()),no=int(str(qingdan['xuhao'])),designBqs=qingdan['nums'],buildBqs=qingdan['jsnums'],BQItemName=qingdan['name'],BQItemUnit=qingdan['unit'],allunitrate=qingdan['price1'],allrate=qingdan['price2'],ispriceLocked=0,insert_dt=curr_time,pact_id=pactid,finishBqs=0)
                else :
                    currqd = {}
                    currtreenode = {}
                    currqd['qingdanid'] = qingdan.get('code','孔')
                    currqd['code'] = cell[1].value
                    currqd['name'] = cell[2].value
                    currqd['unit'] = cell[3].value
                    quantites = str(cell[4].value).strip()
                    if quantites == None:
                        quantites = '0'
                    currqd['nums'] = float(quantites)
                    unitprice = str(cell[5].value).strip()
                    totalprice = str(cell[7].value).strip()
                    if unitprice==None or unitprice=='None':
                        unitprice = '0'
                    if totalprice==None or totalprice=='None':
                        totalprice = '0'
                    currqd['price1'] = Decimal(float(unitprice)).quantize(Decimal('0.00'))
                    currqd['price2'] = Decimal(float(totalprice)).quantize(Decimal('0.00'))
                    currqd['jsnums'] = 0
                    currqd['jsprice'] = 0
                    currlistqd.append(currqd)
                    itemcode = currqd['code'].strip()
                    ##规则匹配
                    if RtItem.objects.filter(RQItem_Code=str(currqd['code'].strip()),BQ_Item_id=currbqitem.id).count()<=0:
                        RqItemmodel = RtItem.objects.create(RQItem_Code=str(currqd['code'].strip()),RT_Item_Name=currqd['name'],unit=currqd['unit'],BQ_Item_id=currbqitem.id,quantities=currqd['nums'],unitprice=currqd['price1'],totalprice=currqd['price2'])
                        rtresourcerelate(itemcode,RqItemmodel.id,RqItemmodel)
        #生成费率表，金额是分部分项工程表中的 施工价格项之和（施工价格=sum(施工工程量×综合单价)，施工工程量是通过四建
        pactid = str(pactid)
        operate = str(operate)
        PactSpaceRate.objects.create(pact_id=pactid,num=1,name='分部分项合计',money=0,calc_method='ratemethod("1","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(pact_id=pactid,num=2,name='措施项目合计',money=0,calc_method='ratemethod("2","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(pact_id=pactid,num=3,name='其他项目合计',money=0,calc_method='ratemethod("3","'+pactid+'","'+operate+'")')
        guifeiobj = PactSpaceRate.objects.create(pact_id=pactid,num=4,name='规费',money=0,canadd=1,calc_method='ratemethod("4","'+pactid+'","'+operate+'")')
        shbaoxianfei = PactSpaceRate.objects.create(canadd=1,num='4.1',parent_id=guifeiobj.id,name=u'社会保险费',pact_id=pactid,calc_method='ratemethod("4.1","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(canadd=0,num='4.1.1',parent_id=shbaoxianfei.id,name=u'管理人员部分',pact_id=pactid,rate=5.38,calc_method='ratemethod("4.1.1","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(canadd=0,num='4.1.2',parent_id=shbaoxianfei.id,name=u'生产工人部分',pact_id=pactid,rate=33.04,calc_method='ratemethod("4.1.2","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(canadd=0,num='4.2',parent_id=guifeiobj.id,name=u'住房公积金',pact_id=pactid,rate=1.96,calc_method='ratemethod("4.2","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(pact_id=pactid,num=5,name='增值税',money=0,rate=11,calc_method='ratemethod("5","'+pactid+'","'+operate+'")')
        PactSpaceRate.objects.create(pact_id=pactid,num=6,name='工程造价',money=0,calc_method='ratemethod("6","'+pactid+'","'+operate+'")')

    elif operate==3:
        excutesqls = []
        for index,cell in enumerate(row):
            if index>1 and cell[0].value != None:
                print u''+str(cell[4].value) + '---' +str(cell[1].value)
                excutesqls.append(OtherItem(pact_id=pactid,insert_dt=curr_time,no=cell[0].value,other_item_name=cell[1].value,other_item_unit=cell[2].value,design_bqs=dataformat(cell[3].value),alunitrate=dataformat(cell[4].value),allrate=dataformat(cell[5].value),percent_labour=dataformat(cell[6].value)))
        if len(excutesqls)>0:
            OtherItem.objects.bulk_create(excutesqls)
        #生成费率表
        pactid = str(pactid)
        operate = str(operate)
        totalmoney = OtherItem.objects.filter(pact_id=pactid).aggregate(total_money=Sum('allrate'))['total_money']
        OtherItemRate.objects.create(pact_id=pactid,num=1,name='分部分项合计',money=0,calc_method='ratemethod("1","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(pact_id=pactid,num=2,name='措施项目合计',money=0,calc_method='ratemethod("2","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(pact_id=pactid,num=3,name='其他项目合计',money=totalmoney,calc_method='ratemethod("3","'+pactid+'","'+operate+'")')
        guifeiobj = OtherItemRate.objects.create(pact_id=pactid,num=4,name='规费',money=0,canadd=1,calc_method='ratemethod("4","'+pactid+'","'+operate+'")')
        shbaoxianfei = OtherItemRate.objects.create(canadd=1,num='4.1',parent_id=guifeiobj.id,name=u'社会保险费',pact_id=pactid,calc_method='ratemethod("4.1","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(canadd=0,num='4.1.1',parent_id=shbaoxianfei.id,name=u'管理人员部分',pact_id=pactid,rate=5.38,calc_method='ratemethod("4.1.1","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(canadd=0,num='4.1.2',parent_id=shbaoxianfei.id,name=u'生产工人部分',pact_id=pactid,rate=33.04,calc_method='ratemethod("4.1.2","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(canadd=0,num='4.2',parent_id=guifeiobj.id,name=u'住房公积金',pact_id=pactid,rate=1.96,calc_method='ratemethod("4.2","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(pact_id=pactid,num=5,name='增值税',money=0,rate=11,calc_method='ratemethod("5","'+pactid+'","'+operate+'")')
        OtherItemRate.objects.create(pact_id=pactid,num=6,name='工程造价',money=0,calc_method='ratemethod("6","'+pactid+'","'+operate+'")')

    return True



'''定额资源匹配'''
def rtresourcerelate(itemno,rtid,rtitemmodel):
    itemno_new = itemno
    if u'换' in itemno:
        itemno_new = itemno.replace(u'换','')
        if KnowledgeRtItem.objects.filter(ItemNo=itemno_new).count()>0:
            # rtitem = KnowledgeRtItem.objects.filter(ItemNo=itemno_new)[0]
            for rtitem in KnowledgeRtItem.objects.filter(ItemNo=itemno_new):
                if u'sz' not in rtitem.chapter.RT_Code_No:
                    addresourcerelate(rtitem,rtid,rtitemmodel.BQ_Item.pact.id)
    elif u'市' in itemno:
        itemno_new = itemno.replace(u'市','')
        if KnowledgeRtItem.objects.filter(ItemNo=itemno_new).count()>0:
            rtitemlist = KnowledgeRtItem.objects.filter(ItemNo=itemno_new)
            for rtitem in rtitemlist:
                if KnowledgeRtChapter.objects.filter(id=rtitem.chapter.id,RT_Code_No__icontains='sz').count()>0:
                    print u'添加市政定额关系'
                    addresourcerelate(rtitem,rtid,rtitemmodel.BQ_Item.pact.id)
        else:
            print u'序号'+itemno+'没有匹配到市政定额数据'
    elif u'安' in itemno:
        itemno_new = itemno.replace(u'安','')
        if KnowledgeRtItem.objects.filter(ItemNo=itemno_new).count()>0:
            rtitemlist = KnowledgeRtItem.objects.filter(ItemNo=itemno_new)
            for rtitem in rtitemlist:
                if KnowledgeRtChapter.objects.filter(id=rtitem.chapter.id,RT_Code_No__icontains='az').count()>0:
                    print u'添加安装定额关系'
                    addresourcerelate(rtitem,rtid,rtitemmodel.BQ_Item.pact.id)
        else:
            print u'序号'+itemno+'没有匹配到安装定额数据'
    else:
        print u'itemno_new-->'+itemno_new
        if KnowledgeRtItem.objects.filter(ItemNo=itemno_new).count()>0:
            # print u'序号'+itemno+'匹配++++++++'
            for rtitem in KnowledgeRtItem.objects.filter(ItemNo=itemno_new):
                if u'sz' not in rtitem.chapter.RT_Code_No:
                    addresourcerelate(rtitem,rtid,rtitemmodel.BQ_Item.pact.id)
        else:
            print u'添加到清单材料表中'
            if KnowledgeRtItemResource.objects.filter(ItemCode=itemno_new).count()>0:
                for knowledgeResource in KnowledgeRtItemResource.objects.filter(ItemCode=itemno_new):
                    if knowledgeResource.price == None:
                        knowledgeResource.price = 0
                    sourcetype = 1
                    if knowledgeResource.restype=='machine':
                        sourcetype = 3
                    elif knowledgeResource.restype=='material':
                        sourcetype = 2
                    else:
                        sourcetype = 1
                    if knowledgeResource.unit_id == None:
                        unititem = ''
                    else:
                        unititem =knowledgeResource.unit.name
                    if RtItemResource.objects.filter(Code=knowledgeResource.ItemCode).count()<=0:
                        itresource = RtItemResource.objects.create(resourcename=knowledgeResource.name,Code=knowledgeResource.ItemCode,unit=unititem,Price=knowledgeResource.price,type=sourcetype)
                    else:
                        itresource = RtItemResource.objects.filter(Code=knowledgeResource.ItemCode)[0]

                    BqItemResourceRelate.objects.create(bqitem_id=rtitemmodel.BQ_Item.id,resource_id=itresource.id,BQTActualAmount=rtitemmodel.quantities,resource_type=sourcetype)

            else:
                itresource = RtItemResource.objects.create(resourcename=rtitemmodel.RT_Item_Name,Code=str(rtitemmodel.RQItem_Code)+str(rtitemmodel.id),unit=rtitemmodel.unit,Price=rtitemmodel.unitprice,type=2)

                BqItemResourceRelate.objects.create(bqitem_id=rtitemmodel.BQ_Item.id,resource_id=itresource.id,BQTActualAmount=rtitemmodel.quantities,resource_type=2)#未知资源下默认是材料类型
    return True

'''人材机和定额关系'''
def addresourcerelate(rtitem,rtid,pact_id):
    if KnowledgeRtItemRelate.objects.filter(item_id=rtitem.id).count()>0:
        for rtitemrelate in KnowledgeRtItemRelate.objects.filter(item_id=rtitem.id):
            if rtitemrelate.resource.price == None or rtitemrelate.resource.price == 'None':
                rtitemrelate.resource.price = 0
            sourcetype = 1
            if rtitemrelate.resource.restype=='machine':
                sourcetype = 3
            elif rtitemrelate.resource.restype=='material':
                sourcetype = 2
            else:
                sourcetype = 1
            if rtitemrelate.resource.unit_id == None:
                unititem = ''
            else:
                unititem = rtitemrelate.resource.unit.name
            if RtItemResource.objects.filter(Code=rtitemrelate.resource.ItemCode).count()<=0:
                rqitemresource = RtItemResource.objects.create(resourcename=rtitemrelate.resource.name,Code=rtitemrelate.resource.ItemCode,unit=unititem,Price=rtitemrelate.resource.price,type=sourcetype)

            else:
                rqitemresource = RtItemResource.objects.filter(Code=rtitemrelate.resource.ItemCode)[0]

            rtitemobj = RtItem.objects.get(id=rtid)
            rtc = rtitemobj.quantities*rtitemrelate.amount
            RtItemResourceRelate.objects.create(rtitem_id=rtid,amount=rtc,resource_id=rqitemresource.id,RTContentAmount=rtitemrelate.amount,RTActualAmount=rtitemrelate.amount,type=rqitemresource.type)
            RtItemResourceThoundRelate.objects.create(rtitem_id=rtid,amount=rtc,resource_id=rqitemresource.id,RTContentAmount=rtitemrelate.amount,RTActualAmount=rtitemrelate.amount,type=rqitemresource.type)

'''解析专业分包execl'''
def doanalysis_fenbaoexecl(pactid,docid):
    response_data = {}
    docobj = Document.objects.get(id = docid)
    dicitem = docobj.docdirectory.all()[0]
    path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
    wb = load_workbook(path)
    sheetname = wb.get_sheet_names()[0]
    sheetContent = wb[sheetname]
    response_data['sheetname'] = sheetname
    response_data['cc'] = sheetContent.cell(row=3,column=2).value
    row = sheetContent.rows
    currxuhao = ''
    queryList = []
    for index,cell in enumerate(row):
        if index>=1:
            if cell[0].value != None:
                pricestr = cell[5].value
                if pricestr==None or pricestr == 'None':
                    pricestr = '0'
                price = Decimal(float(pricestr)).quantize(Decimal('0.00'))
                isphysical = 0

                if (str(cell[11].value)).strip() == '实物量':
                    isphysical = 1
                if type(cell[0].value) is types.LongType:
                    queryList.append(PactFenbao(isphysical=isphysical,pact_id=pactid,designBqs=cell[4].value,parent_code=currxuhao,name=cell[1].value,code=cell[0].value,unit=cell[2].value,price=price))
                else:
                    currxuhao = str(cell[0].value).strip()
                    queryList.append(PactFenbao(isphysical=isphysical,pact_id=pactid,designBqs=cell[4].value,parent_code='',name=cell[1].value,code=currxuhao,unit=cell[2].value,price=price))

    PactFenbao.objects.bulk_create(queryList)

'''解析劳务分包execl'''
def doanalysis_laowuexecl(pactid,docid):
    response_data = {}
    docobj = Document.objects.get(id = docid)
    dicitem = docobj.docdirectory.all()[0]
    path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
    wb = load_workbook(path)
    sheetname = wb.get_sheet_names()[0]
    sheetContent = wb[sheetname]
    row = sheetContent.rows
    currxuhao = ''
    queryList = []
    for index,cell in enumerate(row):
        if index>=1:
            name = cell[1].value
            unit = cell[2].value
            tax = cell[3].value
            taxprice = cell[4].value
            isphysicalstr = str(cell[5].value)
            isphysical = 0
            if (str(cell[5].value)).strip() == '实物量':
                isphysical = 1
            if cell[0].value != None:
                queryList.append(PactLabour(pact_id=pactid,name=name,unit=unit,taxprice=taxprice,tax=tax,isphysical=isphysical))
    PactLabour.objects.bulk_create(queryList)


'''费率计算接口通用type-预算类型（1-分部分项，2-措施项目合计，3-其他项目合计） num-序号，pactid-对应的预算id'''
def ratemethod(num,pactid,type):
    # print type,pactid,num
    if type=='1' or type=='2':
        pr = PactSpaceRate.objects.filter(pact_id=pactid,num=num)[0]
        print pr.name
        if num=='1':
            if type==1:
                pr.money = BqItem.filter(pact_id=pactid,pact__type=item.num).aggregate(sum_money=Sum(F('buildBqs')*F('allunitrate'), output_field=FloatField()))['sum_money']
            else:
                pr.money=0
        elif num=='2':
            if type==2:
                pr.money = BqItem.filter(pact_id=pactid,pact__type=item.num).aggregate(sum_money=Sum(F('buildBqs')*F('allunitrate'), output_field=FloatField()))['sum_money']
            else:
                pr.money=0
        elif num=='3':
            pr.money=0
        elif num=='4':
            print u'不计算'
            pr.money=0
        elif num=='4.1':
            print u'稍后计算'
            pr.money=0
        elif len(num.split('.'))==2 and num !='4.1':
            rengongmoney = RtItemResourceRelate.objects.filter(rtitem__BQ_Item__pact_id=pactid,type=1).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
            rengongmoney = 0 if rengongmoney==None else rengongmoney
            pr.money = pr.rate*rengongmoney/100
            parentrate = PactSpaceRate.objects.get(id=pr.parent_id)
            parentrate.money = parentrate.money+pr.money
            parentrate.save()
        elif len(num.split('.'))==3:
            rengongmoney = RtItemResourceRelate.objects.filter(rtitem__BQ_Item__pact_id=pactid,type=1).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
            rengongmoney = 0 if rengongmoney==None else rengongmoney
            pr.money = pr.rate*rengongmoney/100
            parentrate = PactSpaceRate.objects.get(id=pr.parent_id)
            parentrate.money = parentrate.money+pr.money
            parentrate.save()
        elif num=='5':
            fenburate = PactSpaceRate.objects.filter(pact_id=pactid,num=1)[0]
            guifeirate = PactSpaceRate.objects.filter(pact_id=pactid,num=4)[0]
            pr.money = (guifeirate.money+fenburate.money)*pr.rate/100
        elif num=='6':
            pr.money = PactSpaceRate.objects.filter(Q(pact_id=pactid)&Q(parent_id=0)&~Q(num='6')).aggregate(sum_money=Sum('money'))['sum_money']
        pr.save()
        return pr
    elif type=='3':
        pr = OtherItemRate.objects.filter(pact_id=pactid,num=num)[0]
        if num=='1' or num=='2':
            pr.money = 0
        elif num=='3':
            pr.money = OtherItem.objects.filter(pact_id=pactid).aggregate(sum_money=Sum('allrate'))['sum_money']
        elif num=='4':
            print u'不计算'
            pr.money=0
        elif num=='4.1':
            print u'稍后计算'
            pr.money=0
        elif len(num.split('.'))==2 and num !='4.1':
            rengongmoney = OtherItem.objects.filter(pact_id=pactid).aggregate(sum_money=Sum(F('percent_labour')*F('allrate'), output_field=FloatField()))['sum_money']
            rengongmoney = 0 if rengongmoney==None else rengongmoney
            pr.money = pr.rate*rengongmoney/100
            parentrate = OtherItemRate.objects.get(id=pr.parent_id)
            parentrate.money = parentrate.money+pr.money
            parentrate.save()
        elif len(num.split('.'))==3:
            rengongmoney = OtherItem.objects.filter(pact_id=pactid).aggregate(sum_money=Sum(F('percent_labour')*F('allrate'), output_field=FloatField()))['sum_money']
            rengongmoney = 0 if rengongmoney==None else rengongmoney
            pr.money = pr.rate*rengongmoney/100
            parentrate = OtherItemRate.objects.get(id=pr.parent_id)
            parentrate.money = parentrate.money+pr.money
            parentrate.save()
        elif num=='5':
            qthjrate = OtherItemRate.objects.filter(pact_id=pactid,num=1)[0]
            guifeirate = OtherItemRate.objects.filter(pact_id=pactid,num=4)[0]
            pr.money = (guifeirate.money+qthjrate.money)*pr.rate/100
        elif num=='6':
            pr.money = OtherItemRate.objects.filter(Q(pact_id=pactid)&Q(parent_id=0)&~Q(num='6')).aggregate(sum_money=Sum('money'))['sum_money']
        pr.save()
        return pr
