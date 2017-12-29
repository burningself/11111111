# -*- coding: utf-8 -*-
import traceback,json,sys,datetime,types,random,time,urllib2,urllib,cookielib,requests
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import get_template
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render,render_to_response
from django.template import loader,Context,RequestContext
from Business.models import *
from TaskAndFlow.models import *
from UserAndPrj.models import *
from django.db.models import Q
from django.core import serializers
from openpyxl import Workbook
from openpyxl import load_workbook
from Scc4PM.settings import UPLOAD_DIR
from dss.Serializer import serializer
from django.core.paginator import Paginator
from PIL import Image
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from Scc4PM.settings import CURRENT_PROJECT_ID,DATABASES
from django.db.models import *
from decimal import Decimal
from TaskAndFlow.utility_filemanager import *
from Business.serializers import *
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.authentication import SessionAuthentication
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import renderer_classes
from rest_framework.renderers import StaticHTMLRenderer
import rest_framework_filters
from Business.utils import *

'''总包合同/预算管理开始'''

'''空间管理'''
def spacemanager(request):
    ''' 空间管理 otype 操作标志'''
    response_data = {}
    response_data['issuc'] = True
    try:
        otype = request.GET.get('type')
        if otype == 'add':
            spacename = request.GET.get('spacename')
            curr_time = datetime.datetime.now()
            spaceitem = PactSpace.objects.filter().order_by('-order_no')[0]
            orderno = spaceitem.order_no+1
            PactSpace.objects.create(order_no=orderno,name=spacename,insert_dt=curr_time.strftime('%Y-%m-%d %H:%M:%S'))
        elif otype == 'update':
            spacename = request.GET.get('spacename')
            spaceid = request.GET.get('id')
            if PactSpace.objects.filter(id=spaceid).count()>0:
                spaceitem = PactSpace.objects.get(id=spaceid)
                spaceitem.name = spacename
                spaceitem.save()
        elif otype == 'delete':
            spaceid = request.GET.get('id')
            if PactSpace.objects.filter(id=spaceid).count()>0:
                spaceitem = PactSpace.objects.filter(id=spaceid).delete()
        elif otype == 'move':
            spaceid = request.GET.get('id')
            direction = int(request.GET.get('dir'))
            spaceitem = PactSpace.objects.get(id=spaceid)
            curr_orderno = spaceitem.order_no
            if direction==1:
                if PactSpace.objects.filter(order_no__lt=curr_orderno).count()>0:
                    prespaceitem = PactSpace.objects.filter(order_no__lt=curr_orderno).order_by('-order_no')[0]
                    pre_orderno = prespaceitem.order_no
                    prespaceitem.order_no = curr_orderno
                    spaceitem.order_no = pre_orderno
                    spaceitem.save()
                    prespaceitem.save()
            elif direction==-1:
                if PactSpace.objects.filter(order_no__gt=curr_orderno).count()>0:
                    nextspaceitem = PactSpace.objects.filter(order_no__gt=curr_orderno).order_by('order_no')[0]
                    next_orderno = nextspaceitem.order_no
                    nextspaceitem.order_no = curr_orderno
                    spaceitem.order_no = next_orderno
                    spaceitem.save()
                    nextspaceitem.save()
        else:
            response_data['issuc'] = False
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()

    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''费率表操作'''
@csrf_exempt
def spacefeilv(request):
    response_data = {}
    response_data['issuc'] = True
    try:
        otype = request.GET.get('type')
        if otype == None:
            otype = request.POST.get('type')
        if otype == 'add':
            name = request.GET.get('name')
            pact_id = request.GET.get('pact_id')
            parent_id = int(request.GET.get('parent_id'))
            spacerate = PactSpaceRate.objects.filter(id=parent_id)[0]
            count = PactSpaceRate.objects.filter(parent_id=parent_id).count()+1
            num = spacerate.num+'.'+str(count)
            canadd = 1
            if spacerate.num == '4':
                if (num=='4.2' or num=='4.3'):canadd = 0
            else:
                parent = PactSpaceRate.objects.filter(num='4',pact_id=pact_id)[0]
                if spacerate.parent_id == parent.id:canadd = 0
            pt = Pact.objects.get(id=pact_id)
            PactSpaceRate.objects.create(canadd=canadd,num=num,parent_id=parent_id,name=name,pact_id=pact_id,calc_method='ratemethod("'+num+'","'+str(pact_id)+'","'+str(pt.budgetcont_type)+'")')
        elif otype == 'update':
            feilvid = request.POST.get('pk')
            operate = int(request.POST.get('name'))
            pactspaceitem = PactSpaceRate.objects.get(id=feilvid)
            #修改费率表值 operate 1-修改名称，2-修改费率，3-修改金额
            if operate == 1:
                value = request.POST.get('value')
                response_data['type'] = 1
                pactspaceitem.name = value
            elif operate == 2:
                value = request.POST.get('value')
                response_data['type'] = 2
                pactspaceitem.rate = float(value)
            else:
                value = request.POST.get('value')
                response_data['type'] = 3
                pactspaceitem.money = float(value)
            pactspaceitem.save()
            response_data['value'] = value
            response_data['pact_id'] = pactspaceitem.pact.id
            response_data['id'] = feilvid
        elif otype == 'delete':
            feilvid = request.GET.get('feilvid')
            response_data['pact_id'] = PactSpaceRate.objects.filter(id=feilvid)[0].pact.id
            PactSpaceRate.objects.filter(parent_id=feilvid).delete()
            PactSpaceRate.objects.filter(id=feilvid).delete()
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()

    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''获取费率或者工料机，operate 1-获取费率，2-获取工料机，3-获取2000定额工料机'''
def getpsrate_or_resource(request):
    response_data = {}
    response_data['issuc'] = True
    try:
        pact_id = int(request.GET.get('pact_id'))
        operate = int(request.GET.get('operate',1))
        if operate==1:
            #分部分项合计、措施项目合计 人工费。定义变量，分部分项合计，措施项目合计，其他项目合计，规费，纸质税，人工费
            ratelist = []
            if Pact.objects.get(id=pact_id).budgetcont_type==3:
                for item in OtherItemRate.objects.filter(pact_id=pact_id):
                    eval(item.calc_method)
                #计算其他项目合计费率子父级关系
                for item in OtherItemRate.objects.filter(pact_id=pact_id,parent_id=0):
                    itobj = serializer(item)
                    itobj['child'] = []
                    if item.num=='4':
                        for x in OtherItemRate.objects.filter(parent_id=item.id):
                            xobj = serializer(x)
                            xobj['child'] = []
                            if x.num=='4.1':
                                for y in OtherItemRate.objects.filter(parent_id=x.id):
                                    xobj['child'].append(serializer(y))
                            itobj['child'].append(xobj)
                    ratelist.append(itobj)
            else:
                for item in PactSpaceRate.objects.filter(pact_id=pact_id):
                    eval(item.calc_method)
                #计算分部分项合计费率子父级关系
                for item in PactSpaceRate.objects.filter(pact_id=pact_id,parent_id=0):
                    itobj = serializer(item)
                    itobj['child'] = []
                    if item.num=='4':
                        for x in PactSpaceRate.objects.filter(parent_id=item.id):
                            xobj = serializer(x)
                            xobj['child'] = []
                            if x.num=='4.1':
                                for y in PactSpaceRate.objects.filter(parent_id=x.id):
                                    xobj['child'].append(serializer(y))
                            itobj['child'].append(xobj)
                    ratelist.append(itobj)
            response_data['ratelist'] = ratelist
        elif operate==2:
            resourcelist = []
            qqcs = ComponentQuantities.objects.filter(bqitem__pact_id=pact_id).count()
            quchong = {}
            for item in RtItemResourceRelate.objects.filter(rtitem__BQ_Item__pact_id=pact_id):
                resourceitem = serializer(item)
                resourceitem['id'] = item.id
                if qqcs==0:
                    item.RTActualAmount = 0
                if item.amount==None:
                    item.amount = 0
                resourceitem['amount'] = Decimal(float(item.amount)).quantize(Decimal('0.00'))
                resourceitem['resourcename'] = item.resource.resourcename
                resourceitem['unit'] = item.resource.unit
                resourceitem['Code'] = item.resource.Code
                resourceitem['Price'] = item.resource.Price
                resourceitem['totalprice'] = Decimal(float(item.resource.Price)*float(item.RTContentAmount)).quantize(Decimal('0.00'))
                resourceitem['type'] = item.resource.type
                resourcelist.append(resourceitem)
                if quchong.has_key(resourceitem['Code']):
                    quchong[resourceitem['Code']]['totalprice'] = quchong[resourceitem['Code']]['totalprice']+resourceitem['totalprice']
                    quchong[resourceitem['Code']]['amount'] = quchong[resourceitem['Code']]['amount']+resourceitem['amount']
                else:
                    quchong[resourceitem['Code']] = resourceitem
            response_data['resourcelist'] = serializer(quchong)
        elif operate==3:
            resourcelist = []
            qqcs = ComponentQuantities.objects.filter(bqitem__pact_id=pact_id).count()
            for item in RtItemResourceThoundRelate.objects.filter(rtitem__BQ_Item__pact_id=pact_id):
                resourceitem = serializer(item)
                resourceitem['id'] = item.id
                if qqcs==0:
                    item.RTActualAmount = 0
                if item.amount==None:
                    item.amount = 0
                resourceitem['amount'] = Decimal(float(item.amount)).quantize(Decimal('0.00'))
                resourceitem['resourcename'] = item.resource.resourcename
                resourceitem['unit'] = item.resource.unit
                resourceitem['Code'] = item.resource.Code
                resourceitem['Price'] = item.resource.Price
                resourceitem['totalprice'] = Decimal(float(item.resource.Price)*float(item.RTContentAmount)).quantize(Decimal('0.00'))
                resourceitem['type'] = item.resource.type
                resourcelist.append(resourceitem)
            response_data['resourcelist'] = serializer(resourcelist)
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

def test(request):
    templateName = 'TaskAndFlow/flowtemplate/business_test.html'
    return render_to_response(templateName, RequestContext(request, locals()))
@login_required(login_url="/login/")
def businessmanager(request):
    templateName = 'TaskAndFlow/flowtemplate/businessmanager.html'
    version = time.time()
    try:
        rtitemresources = []
        if RtItemResource.objects.filter().count()>0:
            for resourceitem in RtItemResource.objects.all():
                item = serializer(resourceitem)
                item['description'] = resourceitem.Code+resourceitem.resourcename
                rtitemresources.append(item)
        companys = Company.objects.all()
        professionals = UserMajor.objects.all()
        pacttypes = PactType.objects.all()
    except Exception as e:
        traceback.print_exc()
    return render_to_response(templateName, RequestContext(request, locals()))

'''预算页面数据初始化,获取清单、定额、工料机级联数据'''
def initmanagerData(request):
    response_data = {}
    response_data['issuc'] = True
    try:
        pact_id = int(request.GET.get('pact_id',0))
        bqid = int(request.GET.get('bqid',0))
        deid = int(request.GET.get('deid',0))
        #获取展示的清单信息和定额信息，bqitemlist-清单列表，rtlist-定额列表，rtitem-定额内容,dingarr-工料机列表
        if pact_id == 0:
            if bqid==0:
                bqobj = BqItem.objects.filter().order_by('pact_id')[0]
                bqitemlist = serializer(BqItem.objects.filter(pact_id=bqobj.pact_id))
                response_data['bqitemlist'] = serializer(bqitemlist)
                bqitem = BqItem.objects.filter().order_by('id')[0]
            else:
                bqitem = BqItem.objects.get(id=bqid)
        else:
            #合计项目清单获取
            if Pact.objects.get(id=pact_id).budgetcont_type==3:
                response_data['bqitemlist'] = serializer(OtherItem.objects.filter(pact_id=pact_id))
                return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
            #分部分项合计、措施项目合计数据获取
            if bqid==0:
                bqobj = BqItem.objects.filter(pact_id=pact_id).order_by('pact_id')[0]
                bqitemlist = serializer(BqItem.objects.filter(pact_id=pact_id))
                response_data['bqitemlist'] = serializer(bqitemlist)
                bqitem = BqItem.objects.filter(pact_id=pact_id).order_by('id')[0]
            else:
                bqitemlist = serializer(BqItem.objects.filter(pact_id=pact_id))
                response_data['bqitemlist'] = serializer(bqitemlist)
                bqitem = BqItem.objects.get(id=bqid)
        rtlist = []
        dingarr =[]
        if deid == 0:
            rtitemlist = RtItem.objects.filter(BQ_Item_id=bqitem.id)
        else:
            rtitemlist = RtItem.objects.filter(id=deid)
        for index in range(len(rtitemlist)):
            rtitemmodel = rtitemlist[index]
            rtmodel = serializer(rtitemmodel)
            comparison = 0#comparison用于定额价格与工料机合价的对比
            rtmodel['pricecomparison'] = 1#默认价格一样
            if RtItemResourceRelate.objects.filter(rtitem_id=rtitemmodel.id).count()>0:
                comparison = RtItemResourceRelate.objects.filter(rtitem_id=rtitemmodel.id).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
                ca = comparison+0.05
                cb = comparison-0.05
                comparison = Decimal(comparison).quantize(Decimal('0.00'))
                ca = Decimal(ca).quantize(Decimal('0.00'))
                cb = Decimal(cb).quantize(Decimal('0.00'))
                if ca >= rtitemmodel.unitprice and cb <= rtitemmodel.unitprice:
                    rtmodel['pricecomparison'] = 1
                else:
                    rtmodel['pricecomparison'] = 0
            rtlist.append(rtmodel)
            if index == 0:
                response_data['rtitem'] = serializer(rtmodel)
                rtitem = rtitemmodel
                for dingeitem in RtItemResourceRelate.objects.filter(rtitem_id=rtitem.id).order_by('type'):
                    item = serializer(dingeitem)
                    item['RTItemID'] = dingeitem.rtitem.id
                    item['RQItem_Code'] = dingeitem.rtitem.RQItem_Code
                    item['resourceid'] = dingeitem.resource.id
                    item['resourcetype'] = dingeitem.resource.type
                    item['resourcename'] = dingeitem.resource.resourcename
                    item['resourceunit'] = dingeitem.resource.unit
                    item['resourcecode'] = dingeitem.resource.Code
                    item['resourceprice'] = dingeitem.resource.Price
                    dingarr.append(item)
        response_data['dingarr'] = serializer(dingarr)
        response_data['rtlist'] = serializer(rtlist)
    except Exception as e:
        response_data['issuc'] = False
        response_data['dingarr'] = []
        response_data['rtlist'] = []
        response_data['rtitem'] = {}
        response_data['bqitemlist'] = []
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''预算处理，修改、删除、查询'''
def pacthandle(request):
    response_data = {}
    response_data['issuc'] = True
    try:
        otype = request.GET.get('type')
        if otype == 'info':
            pact_id = request.GET.get('pact_id')
            pactitem = Pact.objects.get(id=pact_id)
            response_data['pactitem'] = serializer(pactitem)
        elif otype == 'update':
            pact_name = request.GET.get('pact_name')
            pact_id = request.GET.get('pact_id')
            pactitem = Pact.objects.get(id=pact_id)
            pactitem.name = pact_name
            pactitem.save()
        elif otype == 'delete':
            pact_id = request.GET.get('pact_id','')
            if pact_id!='':
                pactmodel = Pact.objects.get(id=pact_id)
                if pactmodel.locked == 1:
                    response_data['issuc'] = False
                    response_data['msg'] = u'该预算已经锁定'
                else:
                    pactmodel.delete()

    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''添加预算，解析预算execl'''
def addpact(request):
    response_data = {}
    response_data['issuc'] = True
    response_data['msg'] = u'添加预算成功'
    try:
        space_id = int(request.GET.get('space_id'))
        zbao_isself = int(request.GET.get('zbao_isself'))
        zbao_contenttype = int(request.GET.get('zbao_contenttype'))
        zbao_pactname = request.GET.get('zbao_pactname')
        zbao_pactcode = request.GET.get('zbao_pactcode')
        zbao_pacttype = int(request.GET.get('zbao_pacttype'))
        zbao_professional = int(request.GET.get('zbao_professional'))
        zbao_pactdesc = request.GET.get('zbao_pactdesc')
        uploadfile_qingdanStr = request.GET.get('uploadfile_qingdanStr')
        uploadfile_fujianStr = request.GET.get('uploadfile_fujianStr')
        zbao_company = int(request.GET.get('zbao_company'))
        curr_time = datetime.datetime.now()
        zbao_pactdesc = zbao_pactdesc.replace("\n", "")
        pactmodel = Pact.objects.create(budgetcont_type=zbao_contenttype,is_self=zbao_isself,space_id=space_id,name=zbao_pactname,pactcode=zbao_pactcode,type_id=zbao_pacttype,description=zbao_pactdesc,cooperation_id=zbao_company,hostuser_id=request.user.id,major_id=zbao_professional,insert_dt=curr_time.strftime('%Y-%m-%d %H:%M:%S'))

        if uploadfile_qingdanStr!='':
            docs = uploadfile_qingdanStr
            destdir=None
            if Directory.objects.filter(name='商务管理',islock=True):
                destdir = Directory.objects.get(name='商务管理',islock=True)
            doclist = []
            docsarr = docs.split("#")
            for docid in docsarr:
                if Document.objects.filter(id=docid).count()==1:
                    tar = Document.objects.get(id=docid)
                    if destdir:
                        tar.docdirectory.add(destdir)
                        movefiletoDir(tar,destdir)
                else:
                    raise Exception(u'文件不存在！')
                doclist.append(PactRelatefile(pact_id=pactmodel.id,file_id=docid,file_type=1,pact_type=1))
            PactRelatefile.objects.bulk_create(doclist)
        relatefileobj = PactRelatefile.objects.filter(pact_id=pactmodel.id,file_type=1,pact_type=1)[0]
        #zbao_contenttype 预算类型1:-分部分项，2-措施合计，3-其他项目合计，三者execl格式不同
        execlhandle(relatefileobj.file_id,pactmodel.id,zbao_contenttype)
    except Exception as e:
        print RtItem.objects.filter(BQ_Item__pact_id=pactmodel.id)
        PactRelatefile.objects.filter(pact_id=pactmodel.id).delete()
        if RtItemResourceRelate.objects.filter(rtitem__BQ_Item__pact_id=pactmodel.id).count()>0:
            RtItemResourceRelate.objects.filter(rtitem__BQ_Item__pact_id=pactmodel.id).delete()
        RtItem.objects.filter(BQ_Item__pact_id=pactmodel.id).delete()
        BqItem.objects.filter(pact_id=pactmodel.id).delete()
        Pact.objects.filter(id=pactmodel.id).delete()
        response_data['issuc'] = False
        response_data['msg'] = u'添加预算失败'
        traceback.print_exc()

    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''获取清单信息'''
def getBqitemBycode(request):
    response_data = {}
    response_data['issuc'] = True
    bqcode = request.GET.get('bqcode','')
    if bqcode == '':
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        if BqItem.objects.filter(BQItem_Code=bqcode).count()>0:
            response_data['bqitem'] = serializer(BqItem.objects.filter(BQItem_Code=bqcode)[0])
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'不存在的清单'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''修改定额信息'''
@csrf_exempt
def updateDinge(request):
    response_data = {}
    id = request.POST.get('pk','')
    operate = int(request.POST.get('name',0))
    value = request.POST.get('value','')

    resourcerelate = RtItemResourceRelate.objects.get(id=id)
    if resourcerelate!='':
        if operate == 0:
            resourcerelate.amount = value
        elif operate == 4:
            resourcerelate.RTActualAmount = value
        elif operate == 1:
            resourcerelate.spect = value
        elif operate == 2:
            resourcerelate.amount = float(value)
            resourceitem = RtItemResource.objects.get(id=resourcerelate.resource.id)
            response_data['resourcename'] = resourceitem.resourcename
            response_data['type'] = resourceitem.type
            response_data['resourcecode'] = resourceitem.Code
            response_data['rtid'] = resourcerelate.rtitem.id
            response_data['unit'] = resourceitem.unit
            response_data['sourceid'] = resourceitem.id
            response_data['price'] = str(resourceitem.Price)
        resourcerelate.save()

    comparison = 0      #comparison用于定额价格与工料机合价的对比
    if RtItemResourceRelate.objects.filter(rtitem_id=resourcerelate.rtitem.id).count()>0:
        comparison = RtItemResourceRelate.objects.filter(rtitem_id=resourcerelate.rtitem.id).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
        comparison = Decimal(comparison).quantize(Decimal('0.00'))
    response_data['id']=resourcerelate.rtitem.id
    response_data['relateid']=resourcerelate.id
    response_data['comparison'] = serializer(comparison)
    response_data['issucc'] = 'true'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

#修改清单量时候联动接口
@csrf_exempt
def updateBqitemQuality(request):
    response_data = {}
    response_data['issucc'] = 'true'
    try:
        id = request.POST.get('name','')
        value = request.POST.get('value','')
        bq = BqItem.objects.get(id=id)
        PostUrl = 'http://121.41.22.17:17531/api/PactChanges'
        cookie = cookielib.CookieJar()
        handler = urllib2.HTTPCookieProcessor(cookie)
        opener = urllib2.build_opener(handler)
        postData={
            'bqcode':bq.BQItem_Code,
            'pactid':bq.pact.id,
            'bqamount':value,
            'schema':str(DATABASES['pms']['NAME'])
        }
        headers = {}
        data = urllib.urlencode(postData)
        request = urllib2.Request(PostUrl,data)
        response = opener.open(request)
        result = response.read()
        print result
    except Exception as e:
        traceback.print_exc()
        print u'预算联动失败'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''预算页面左侧树，一步加载'''
def get_business_tree(request):
    try:
        id = request.GET.get('id', '')
        response_data = {}
        child_list = []
        if id == '#':
            treelist = PactSpace.objects.filter().order_by('order_no')
            for each in treelist:
                child_data = {}
                child_data["id"] = 'space-'+str(each.id)
                child_data["text"] = each.name
                child_data["icon"] = "/img/buildings2.png"
                child_data["state"] = {'opened': True}
                child_data_list = []
                children_datazx = {}
                children_datazx["id"] = 'zixing-'+str(each.id)
                children_datazx["text"] = '自行'
                children_datazx["icon"] = "/img/buildings2.png"
                children_datazx["children"] = True
                children_datafb = {}
                children_datafb["id"] = 'fenbao-'+str(each.id)
                children_datafb["text"] = '分包'
                children_datafb["icon"] = "/img/buildings2.png"
                children_datafb["children"] = True
                child_data_list.append(children_datafb)
                child_data_list.append(children_datazx)
                child_data["children"] = child_data_list
                child_list.append(child_data)
        else:
            idarr = id.split('-')
            child_list = []
            if idarr[0] == 'fenbao' or idarr[0] == 'zixing':
                pactlist = []
                node_id = ''
                if idarr[0] == 'fenbao':
                    pactlist = Pact.objects.filter(space_id=idarr[1],is_self=2)
                    node_id = 'zixingyg-'+idarr[1]
                elif idarr[0] == 'zixing':
                    pactlist = Pact.objects.filter(space_id=idarr[1],is_self=1)
                    node_id = 'zixingyg-'+idarr[1]

                for pactitem in pactlist:
                    pactitem_data = {}
                    pactitem_data["id"] = node_id+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/hetong.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = []

                    pactid_str = str(pactitem.id)
                    type_str = str(pactitem.budgetcont_type)
                    fenbufenx = {}
                    fenbufenx["id"] = ('qitaxiangm-' if pactitem.budgetcont_type==3 else 'fenbufenxiang-')+pactid_str+'-'+type_str
                    fenbufenx["text"] = u'其他项目合计' if pactitem.budgetcont_type==3 else u'分部分项工程'
                    fenbufenx["icon"] = "/img/hetong.png"
                    fenbufenx["state"] = {'opened': True}
                    feilvbiao = {}
                    feilvbiao["id"] = ('qtxmflb-' if pactitem.budgetcont_type==3 else 'feilvbiao-')+pactid_str+'-'+type_str
                    feilvbiao["text"] = u'费率表'
                    feilvbiao["icon"] = "/img/hetong.png"
                    feilvbiao["state"] = {'opened': True}
                    cresource = {}
                    cresource["id"] = 'cresource-'+pactid_str+'-'+type_str
                    cresource["text"] = u'工料机表'
                    cresource["icon"] = "/img/hetong.png"
                    cresource["state"] = {'opened': True}
                    cpresource = {}
                    cpresource["id"] = 'cpresource-'+pactid_str+'-'+type_str
                    cpresource["text"] = u'2000定额工料机表'
                    cpresource["icon"] = "/img/hetong.png"
                    cpresource["state"] = {'opened': True}

                    pactitem_data['children'].append(fenbufenx)
                    pactitem_data['children'].append(feilvbiao)
                    if pactitem.budgetcont_type == 1 or pactitem.budgetcont_type == 2:
                        pactitem_data['children'].append(cresource)
                        pactitem_data['children'].append(cpresource)
                    child_list.append(pactitem_data)
        return HttpResponse(json.dumps(child_list), content_type="application/json")
    except:
        traceback.print_exc()

'''工料机处理接口'''
def resourcehandle(request):
    response_data = {}
    response_data['issuc'] = True
    try:
        otype = request.GET.get('type','')
        if otype == 'add':
            code = request.GET.get('code','')
            resourcetype = request.GET.get('resourcetype','1')
            name = request.GET.get('name','')
            unit = request.GET.get('unit','')
            price = float(request.GET.get('price','0'))#Decimal(float(request.GET.get('price','0'))).quantize(Decimal('0.000'))
            if RtItemResource.objects.filter(Code=code).count()>0:
                response_data['issuc'] = False
                response_data['msg'] = '该编号的资源已存在'
            else:
                rqitr = RtItemResource.objects.create(resourcename=name,Code=code,unit=unit,Price=price,type=int(resourcetype))
                response_data['resource_id'] = rqitr.id
        elif otype == 'delete':
            id = request.GET.get('id')
            RTItemID = RtItemResourceRelate.objects.filter(id=id)[0].rtitem.id
            response_data['issuc'] = True
            RtItemResourceRelate.objects.filter(id=id).delete()
            comparison = 0#comparison用于定额价格与工料机合价的对比
            if RtItemResourceRelate.objects.filter(rtitem_id=RTItemID).count()>0:
                comparison = RtItemResourceRelate.objects.filter(rtitem_id=RTItemID).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
                comparison = Decimal(comparison).quantize(Decimal('0.00'))
            response_data['comparison'] = serializer(comparison)
            response_data['RTItemID'] = RTItemID
    except Exception as e:
        response_data['issuc'] = False
        response_data['msg'] = '程序错误'
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''解析人材机execl'''
def analysis_rencaiji_execl(request):
    response_data = {}
    response_data['issuc'] = True
    docid = request.GET.get('docid')
    docobj = Document.objects.get(id = docid)
    dicitem = Directory.objects.get(name='商务管理',islock=True)
    docobj.docdirectory.add(dicitem)
    movefiletoDir(docobj,dicitem)
    dicitem = docobj.docdirectory.all()[0]
    path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
    wb = load_workbook(path,data_only=True)
    sheetname = wb.get_sheet_names()[0]
    try:
        response_data['sheetname'] = sheetname
        sheetContent = wb[sheetname]
        row = sheetContent.rows
        executeList = []
        num = 0
        for index,cell in enumerate(row):
            if index >= 3:
                if cell[0].value != None and cell[0].value != '序号' and cell[1].value != None and cell[2].value != None :
                    num = 0
                    resourcename = (cell[2].value).strip()
                    unit = (str(cell[4].value)).strip()
                    Code = (str(cell[1].value)).strip()
                    decprice = str(cell[6].value).strip()
                    if decprice==None or decprice == 'None':
                        decprice = '0'
                    Price = Decimal(float(decprice)).quantize(Decimal('0.000'))
                    if unit == '工日':
                        resource_type = 1
                    elif unit == '台班':
                        resource_type = 3
                    else:
                        resource_type = 2
                    if RtItemResource.objects.filter(Code=Code).count()<=0:
                        RtItemResource.objects.create(resourcename=resourcename,unit=unit,Code=Code,type=resource_type,Price=Price)
                    else:
                        resourceitem = RtItemResource.objects.filter(Code=Code)[0]
                        resourceitem.Price = Price
                        resourceitem.save()
                else:
                    num = num+1
                    if num>5:
                        break
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''定额修改人材机'''
def rencaijiadd(request):
    response_data = {}
    RTItemID = request.GET.get('RTItemID')
    RTContentAmount = request.GET.get('RTContentAmount')
    show = request.GET.get('show')
    rtitemmodal = RtItem.objects.get(id=RTItemID)
    if u'市' in rtitemmodal.RQItem_Code or u'换' in rtitemmodal.RQItem_Code or u'市' in rtitemmodal.RQItem_Code:
        rencaiji = int(request.GET.get('rencaiji'))
        resourceitem = RtItemResource.objects.get(id=rencaiji)
        rtitem = RtItemResourceRelate.objects.create(RTContentAmount=RTContentAmount,rtitem_id=RTItemID,RTActualAmount=RTContentAmount,type=resourceitem.type,resource_id=resourceitem.id)
        response_data['issuc'] = True
        item = {}
        item['id'] = rtitem.id
        item['RTItemID'] = rtitem.rtitem.id
        item['RTItemCode'] = rtitemmodal.RQItem_Code
        item['spect'] = rtitem.spect
        item['RTContentAmount'] = rtitem.RTContentAmount
        item['RTActualAmount'] = rtitem.RTActualAmount
        item['resourceid'] = rtitem.resource.id
        item['resourcetype'] = rtitem.resource.type
        item['resourcename'] = rtitem.resource.resourcename
        item['resourceunit'] = rtitem.resource.unit
        item['resourcecode'] = rtitem.resource.Code
        item['resourceprice'] = rtitem.resource.Price
        response_data['resourcerelate'] = serializer(item)
        comparison = 0#comparison用于定额价格与工料机合价的对比
        if RtItemResourceRelate.objects.filter(rtitem_id=RTItemID).count()>0:
            comparison = RtItemResourceRelate.objects.filter(rtitem_id=RTItemID).aggregate(sum_money=Sum(F('RTActualAmount')*F('resource__Price'), output_field=FloatField()))['sum_money']
            comparison = Decimal(comparison).quantize(Decimal('0.00'))
        response_data['comparison'] = serializer(comparison)
    else:
        response_data['issuc'] = False
        response_data['msg'] = '清单关联条目无法添加人材机'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''总包合同管理结束'''

'''分包合同管理开始'''
@login_required(login_url="/login/")
def fenbaopact(request):
    templateName = 'TaskAndFlow/flowtemplate/business_fenbaopact.html'
    companys = Company.objects.all()
    professionals = UserMajor.objects.all()
    yusuans = Pact.objects.filter(type_id=1)
    pacttypes = PactType.objects.all()
    return render_to_response(templateName, RequestContext(request, locals()))

'''添加专业分包和劳务分包'''
def addfenpact(request):
    response_data = {}
    zbao_pactname = request.GET.get('zbao_pactname')
    zbao_pactcode = request.GET.get('zbao_pactcode')
    zbao_pacttype = int(request.GET.get('zbao_pacttype'))
    zbao_professional = request.GET.get('zbao_professional')
    zbao_pacttax = request.GET.get('zbao_pacttax')
    zbao_pactdesc = request.GET.get('zbao_pactdesc')
    uploadfile_qingdanStr = request.GET.get('uploadfile_qingdanStr')
    uploadfile_fujianStr = request.GET.get('uploadfile_fujianStr')
    zbao_yusuan = request.GET.get('zbao_yusuan')
    zbao_company = request.GET.get('zbao_company')
    curr_time = datetime.datetime.now()
    response_data['issuc'] = True
    response_data['msg'] = '添加成功'
    try:
        pactmodel = Pact.objects.create(name=zbao_pactname,pactcode=zbao_pactcode,cess=zbao_pacttax,type_id=zbao_pacttype,description=zbao_pactdesc,cooperation_id=zbao_company,hostuser_id=request.user.id,major_id=zbao_professional,insert_dt=curr_time.strftime('%Y-%m-%d %H:%M:%S'))
        print pactmodel.id
        if uploadfile_qingdanStr!='':
            docs = uploadfile_qingdanStr
            destdir=None
            if Directory.objects.filter(name='商务管理',islock=True):
                destdir = Directory.objects.get(name='商务管理',islock=True)
            doclist = []
            docsarr = docs.split("#")
            for docid in docsarr:
                if Document.objects.filter(id=docid).count()==1:
                    tar = Document.objects.get(id=docid)
                    if destdir:
                        tar.docdirectory.add(destdir)
                        movefiletoDir(tar,destdir)
                else:
                    raise Exception(u'文件不存在！')
                doclist.append(PactRelatefile(pact_id=pactmodel.id,file_id=docid,file_type=1,pact_type=2))

                if zbao_pacttype == 5:
                    doanalysis_fenbaoexecl(pactmodel.id,docid)
                elif zbao_pacttype == 2:
                    doanalysis_laowuexecl(pactmodel.id,docid)

            PactRelatefile.objects.bulk_create(doclist)

        # 分包和预算关联关系入库
        yusuanarr = zbao_yusuan.split(',')
        sqllist = []
        for yusuanitem in yusuanarr:
            PactFenbaoYusuan.objects.create(fenbao_id=int(yusuanitem),yusuan_id=pactmodel.id)
            #添加到规则中
            bqlist = BqItem.objects.filter(pact_id=int(yusuanitem))
            for bqitemobj in bqlist:
                if CalRelation.objects.filter(bqItem_id=bqitemobj.id).count()==0:
                    sqllist.append(CalRelation(valueUnit=bqitemobj.BQItemUnit,bqItem_id=bqitemobj.id,name=bqitemobj.BQItemName))
        if len(sqllist)>0:
            CalRelation.objects.bulk_create(sqllist)
    except Exception as e:
        response_data['issuc'] = False
        response_data['msg'] = '添加失败'
        Pact.objects.filter(id=pactmodel.id).delete()
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

@csrf_exempt
def getpactqingdanlist(request):
    response_data = {}
    try:
        pact_id = int(request.POST.get('param[pact_id]','0'))
        pact_type = int(request.POST.get('param[pact_type]','5'))
        draw = int(request.POST.get('draw','1'))
        start = int(request.POST.get('start','0'))
        pagesize = int(request.POST.get('length','10'))
        end = start+pagesize
        if pact_id != 0:
            if pact_type == 5:
                datalist = PactFenbao.objects.filter(Q(pact_id=pact_id)&~Q(code='一'))[start:end]
                le = PactFenbao.objects.filter(Q(pact_id=pact_id)&~Q(code='一')).count()
            elif pact_type == 2:
                datalist = PactLabour.objects.filter(pact_id=pact_id)[start:end]
                le = PactLabour.objects.filter(pact_id=pact_id).count()
        else:
            if Pact.objects.filter(type_id=5).count()>0:
                pactitem = Pact.objects.filter(type_id=5)[0]
                datalist = PactFenbao.objects.filter(Q(pact_id=pactitem.id)&~Q(code='一'))[start:end]
                le = PactFenbao.objects.filter(Q(pact_id=pactitem.id)&~Q(code='一')).count()
            else:
               datalist = {}
               le = 0
        response_data['data'] = serializer(datalist)
        response_data['draw'] = draw
        response_data['recordsTotal'] = le
        response_data['recordsFiltered'] = le
    except Exception as e:
        print e
        response_data['data'] = {}
        response_data['draw'] = 1
        response_data['recordsTotal'] = 0
        response_data['recordsFiltered'] = 0
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''分包管理树'''
def get_fenbaopact_tree(request):
    id = request.GET.get('id', '')
    type_id = request.GET.get('type_id', 5)
    child_list = []
    fbtreelist = Pact.objects.filter(Q(type_id=2) | Q(type_id=5)).order_by('-type_id')#分包合同
    currtype = 0;
    if len(fbtreelist)>0:
        brothers_data = {}
        brothers_children_data = []
        inum = 0
        for each in fbtreelist:
            child_data = {}
            child_data["id"] = str(each.type_id)+'@@'+str(each.id)
            child_data["text"] = each.name
            child_data["icon"] = "/img/hetong.png"
            if currtype == 0:
                currtype = each.type_id
                brothers_data["id"] = each.type_id
                if each.type_id==2:
                    brothers_data["text"] = u'劳务分包合同'
                else:
                    brothers_data["text"] = u'专业分包合同'
                brothers_data["icon"] = "/img/buildings2.png"
                brothers_data["state"] = {'opened': True}
            if currtype!=each.type_id:
                currtype = each.type_id
                brothers_data['children'] =brothers_children_data
                child_list.append(brothers_data)
                brothers_children_data = []
                brothers_data = {}
                brothers_data["id"] = each.type_id
                if each.type_id==2:
                    brothers_data["text"] = u'劳务分包合同'
                else:
                    brothers_data["text"] = u'专业分包合同'
                brothers_data["icon"] = "/img/buildings2.png"
                brothers_data["state"] = {'opened': True}

                if inum+1==len(fbtreelist):
                    brothers_children_data.append(child_data)
                    brothers_data['children'] =brothers_children_data
                    child_list.append(brothers_data)
                else:
                    brothers_children_data.append(child_data)
            else :
                if inum+1==len(fbtreelist):
                    brothers_children_data.append(child_data)
                    brothers_data['children'] =brothers_children_data
                    child_list.append(brothers_data)
                else:
                    brothers_children_data.append(child_data)
            inum = inum+1
    return HttpResponse(json.dumps(child_list,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''删除分包'''
def delFenbaoitem(request):
    response_data = {}
    response_data['issuc'] = True
    pactid = request.GET.get('pactid','')
    if pactid!='':
        PactRelatefile.objects.filter(pact_id=pactid).delete()
        if PactFenbao.objects.filter(pact_id=pactid).count()>0:
            PactFenbao.objects.filter(pact_id=pactid).delete()
        if PactLabour.objects.filter(pact_id=pactid).count()>0:
            PactLabour.objects.filter(pact_id=pactid).delete()
        Pact.objects.filter(id=pactid).delete()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''分包合同管理结束'''

'''施工预算管理开始'''

'''工程量核算管理主页 status 0-工程量规则，1-工程量汇总，3-revit导出量'''
@login_required(login_url="/login/")
def shigongyusuan(request):
    version = time.time()
    templateName = 'TaskAndFlow/flowtemplate/business_shigongyusuan.html'
    status = int(request.GET.get('status', 1))
    srcmodels = SrcmodelFile.objects.all()
    bqlist = BqItem.objects.filter().order_by('id')
    sclist = PactFenbao.objects.filter()
    calfile = CalRelation.objects.all().count()
    yusuanpacts = Pact.objects.filter(type_id=1)

    #获取专业和劳务分包项，groups用于界面选择代码项的修改
    groups = []
    if PactFenbao.objects.filter().count()>0:
        fenbao_pacts = PactFenbao.objects.values('pact_id').annotate(sum_aa=Max('name'))
        for item in fenbao_pacts:
            fbs = PactFenbao.objects.filter(pact_id=item['pact_id']).exclude(code__in=['一','二','三']).order_by('pact_id')
            obj = {}
            obj['text'] = fbs[0].pact.name
            obj['children'] = []
            for o in fbs:
                oy = {}
                oy['id'] = 'fenbao-'+str(o.id)
                oy['text'] = o.name
                obj['children'].append(oy)
            groups.append(obj)

    if PactLabour.objects.filter().count()>0:
        labour_pacts = PactLabour.objects.values('pact_id').annotate(sum_aa=Max('name'))
        for item in labour_pacts:
            fbs = PactLabour.objects.filter(pact_id=item['pact_id']).order_by('pact_id')
            obj = {}
            obj['text'] = fbs[0].pact.name
            obj['children'] = []
            for o in fbs:
                oy = {}
                oy['id'] = 'labour-'+str(o.id)
                oy['text'] = o.name
                obj['children'].append(oy)
            groups.append(obj)

    #liexianglist-根据CalRelation计算规则表、ComponentQuantities表组装展示数据
    liexianglist = []
    if status!=0:
        dlist = CalRelation.objects.filter().order_by('category','toTaskOrder')
        currcategory = ''#当前分类，根据category字段值不为空时判断，对应界面“列项”
        currdetail = ''#当前类型名称，对应界面“类型”
        currcalcitem = {}#将数据按照类型分组后得到的对象
        childrenlist = []#currcalcitem中的子集
        currnum = 1#dlist索引，辅助用
        totalcount = CalRelation.objects.filter().count()
        for item in dlist:
            currmodel = serializer(item)
            currmodel_B = serializer(item)
            currmodel['bqitem_id'] = currmodel_B['bqitem_id'] = item.bqItem.id
            if ComponentQuantities.objects.filter(calcRelation_id=item.id).count()>0:
                value = ComponentQuantities.objects.filter(calcRelation_id=item.id).aggregate(can_update=Sum('can_update'),valueQuantity = Sum('valueQuantity'),costQuantity = Sum('costQuantity'))
                currmodel['workValueBqs'] = value['valueQuantity']
                currmodel_B['workValueBqs'] = value['costQuantity']
                currmodel['canupdate'] = currmodel_B['canupdate'] = value['can_update']
                if item.bqItem != None:
                    currmodel['totalmoney'] = Decimal(value['valueQuantity']*float(item.bqItem.allunitrate)).quantize(Decimal('0.00'))
                else:
                    currmodel['totalmoney'] = ''
                if item.SCItem != None:
                    currmodel_B['totalmoney'] = Decimal(value['valueQuantity']*float(item.SCItem.price)).quantize(Decimal('0.00'))
                else:
                    currmodel_B['totalmoney'] = ''
            else:
                currmodel['workValueBqs'] = ''
                currmodel_B['workCostBqs'] = ''
                currmodel['canupdate'] = currmodel_B['canupdate'] = 1

            #####根据bqItem、SCItem、lpItem区分产值、分建成本、任务单
            if item.bqItem != None:
                currmodel['BQItem_Code'] = item.bqItem.BQItem_Code
                currmodel['namedesc'] = u'产值'
                currmodel['designBqs'] = item.bqItem.designBqs
                currmodel['price'] = Decimal(item.bqItem.allunitrate).quantize(Decimal('0.00'))
            if item.SCItem != None:
                currmodel_B['namedesc'] = u'分建成本'
                currmodel_B['SCItem_name'] = item.SCItem.name
                currmodel_B['designBqs'] = item.SCItem.designBqs
                currmodel_B['price'] = Decimal(item.SCItem.price).quantize(Decimal('0.00'))
            if item.lpItem != None:
                currmodel_B['namedesc'] = u'任务单'
                currmodel_B['SCItem_name'] = item.lpItem.name
                currmodel_B['designBqs'] = None
                currmodel_B['price'] = Decimal(item.lpItem.taxprice).quantize(Decimal('0.00'))

            if currdetail != item.name:
                currdetail = item.name
                currmodel['detail'] = item.name
                currmodel_B['detail'] = None
            else:
                currmodel['detail'] = currmodel_B['detail'] = None

            if item.category.strip() == '' or currcategory != item.category:
                if currnum == 1:
                    childrenlist.append(currmodel)
                    childrenlist.append(currmodel_B)
                    currcalcitem['category'] = item.category
                    currcalcitem['id'] = item.id
                    currcalcitem['colsrow'] = 2
                elif currnum == totalcount:
                    liexianglist.append(currcalcitem)
                    currcalcitem = {}
                    childrenlist = []
                    currcalcitem['category'] = item.category
                    currcalcitem['id'] = item.id
                    currcalcitem['colsrow'] = 2
                    childrenlist.append(currmodel)
                    childrenlist.append(currmodel_B)
                    currcalcitem['list'] = childrenlist
                    liexianglist.append(currcalcitem)
                else:
                    currcalcitem['list'] = childrenlist
                    liexianglist.append(currcalcitem)
                    currcalcitem = {}
                    childrenlist = []
                    currcalcitem['category'] = item.category
                    currcalcitem['id'] = item.id
                    currcalcitem['colsrow'] = 2
                    currcalcitem['list'] = childrenlist
                    childrenlist.append(currmodel)
                    childrenlist.append(currmodel_B)
                currcategory = item.category

            else:
                # print u'同种分类'
                childrenlist.append(currmodel)
                childrenlist.append(currmodel_B)
                currcalcitem['colsrow'] += 2
                if currnum == totalcount:
                    liexianglist.append(currcalcitem)
            currnum = currnum+1
    return render_to_response(templateName, RequestContext(request, locals()))
#修改规则信息
@csrf_exempt
def updateYusuanRule(request):
    response_data = {}
    response_data['issucc'] = True
    id = int(request.POST.get('pk','0'))
    operate = request.POST.get('name','')
    value = request.POST.get('value','')
    bqitemid = request.POST.get('bqitem_id','')

    if operate == 'addrule':
        unit = request.POST.get('unit','')
        bqobj = BqItem.objects.get(id=bqitemid)
        CalRelation.objects.create(valueUnit=unit,bqItem_id=bqitemid,name=bqobj.BQItemName)
    else:
        calcitem = CalRelation.objects.get(id=int(id))
        if calcitem!='':
            if operate == 'liex':
                calcitem.category = value
                calcitem.save()
            elif operate == 'type':
                calcitem.name = value
                calcitem.save()
            elif operate == 'shuxing':
                calcitem.differParam = value
                calcitem.save()
            elif operate == 'fenbao':
                if CalRelation.objects.filter(SCItem_id=value,bqItem_id=bqitemid).count()==0:
                    scitemobj = PactFenbao.objects.get(id=value)
                    calcitem.lpItem = None
                    calcitem.SCItem = scitemobj
                    calcitem.costUnit = scitemobj.unit
                    calcitem.save()
                else:
                    response_data['issucc'] = False
                    response_data['msg'] = u'已存在该分建成本'
            elif operate == 'labour':
                if CalRelation.objects.filter(lpItem_id=value,bqItem_id=bqitemid).count()==0:
                    lpItemobj = PactLabour.objects.get(id=value)
                    calcitem.SCItem = None
                    calcitem.lpItem = lpItemobj
                    calcitem.costUnit = lpItemobj.unit
                    calcitem.save()
                else:
                    response_data['issucc'] = False
                    response_data['msg'] = u'已存在该任务单'
            response_data['id']=calcitem.id

    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

#自动生成对比规则
def autobudget(request):
    response_data = {}
    try:
        response_data['issuc'] = True
        bqitemlist = BqItem.objects.filter()
        response_data['bqitemlist'] = serializer(bqitemlist)
        for bqitemobj in bqitemlist:
            PactFenbao.objects.filter()
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def addbudget_construct(request):
    response_data = {}
    name = request.GET.get('name')
    srcmodel_id = request.GET.get('scrmodel_id')
    description = request.GET.get('description')
    docid = request.GET.get('docid')
    response_data['issuc'] = True
    try:
        # print u'srcmodel_id=='+str(srcmodel_id)
        budgetmodel = Budget.objects.create(name=name,description=description,doc_id=docid,srcmodel_id=srcmodel_id)
        if docid!='':
            destdir = None
            if Directory.objects.filter(name='商务管理',islock=True):
                destdir = Directory.objects.get(name='商务管理',islock=True)
            if Document.objects.filter(id=docid).count()==1:
                tar = Document.objects.get(id=docid)
                response_data['filename'] = tar.name
                response_data['filepath'] = str(tar.filepath)+str(tar.name)
                if destdir:
                    tar.docdirectory.add(destdir)
                    movefiletoDir(tar,destdir)
            else:
                raise Exception(u'文件不存在！')
            response_data['issuc'] = doanalysis_budgetconstruct(budgetmodel.id,docid)
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def deleteBudget(request):
    response_data = {}
    budget_id = request.GET.get('budget_id','')
    response_data['issuc'] = True
    try:
        if budget_id == '' or Budget.objects.filter(id=int(budget_id)).count()==0:
            response_data['issuc'] = False
            response_data['msg'] = '参数异常或者施工不存在'
        else:
            ComponentQuantities.objects.filter(budget_id=int(budget_id)).delete()
            Budget.objects.filter(id=int(budget_id)).delete()
    except Exception as e:
        response_data['issuc'] = False
        response_data['msg'] = '系统异常'
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def doanalysis_budgetconstruct(budget_id,docid):
    flag = True
    try:
        docobj = Document.objects.get(id = docid)
        dicitem = docobj.docdirectory.all()[0]
        path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
        wb = load_workbook(path)
        sheetname = wb.get_sheet_names()[0]
        sheetContent = wb[sheetname]
        row = sheetContent.rows
        queryList = []
        for index,cell in enumerate(row):
            if (index>=1) and (cell[0].value != None):
                name = str(cell[1].value)
                pro_name = name[:name.find('1.')]
                pro_code = cell[0].value
                amountstr = str(cell[2].value).strip()
                if amountstr==None or amountstr=='' or amountstr=='None':
                    amountstr = '0'
                pro_amount = float(amountstr)
                group_code = cell[4].value
                annotation = name[name.find('1.'):]
                unit = cell[3].value
                # print index
                queryList.append(ComponentQuantities(valueQuantity=pro_amount,name=pro_name,code=pro_code,description=annotation,budget_id=budget_id))

        ComponentQuantities.objects.bulk_create(queryList)
        PostUrl = 'http://121.41.22.17:17531/api/ComponentBQ'
        cookie = cookielib.CookieJar()
        handler = urllib2.HTTPCookieProcessor(cookie)
        opener = urllib2.build_opener(handler)
        postData={
            'budget_id':budget_id,
            'schema':str(DATABASES['pms']['NAME'])
        }
        headers = {}
        data = urllib.urlencode(postData)
        request = urllib2.Request(PostUrl,data)
        response = opener.open(request)
        result = response.read()
    except Exception as e:
        print u'施工预算调用接口异常'
        ComponentQuantities.objects.filter(budget_id=budget_id).delete()
        Budget.objects.filter(id=budget_id).delete()
        Document.objects.get(id=docid).delete()
        flag = False
        traceback.print_exc()
    return flag
def addbudget_requirement(request):
    response_data = {}
    docid = request.GET.get('docid')
    response_data['issuc'] = True
    try:
        if docid!='':
            destdir = None
            if Directory.objects.filter(name='商务管理',islock=True):
                destdir = Directory.objects.get(name='商务管理',islock=True)
            if Document.objects.filter(id=docid).count()==1:
                tar = Document.objects.get(id=docid)
                response_data['filename'] = tar.name
                response_data['filepath'] = str(tar.filepath)+str(tar.name)
                if destdir:
                    tar.docdirectory.add(destdir)
                    movefiletoDir(tar,destdir)
            else:
                raise Exception(u'文件不存在！')
            doanalysis_budgetrequirement(docid)
            # if PactLabour.objects.filter().count()>0 and PactFenbao.objects.filter().count()>0:
            #     doanalysis_budgetrequirement(docid)
            # else:
            #     response_data['issuc'] = False
            #     response_data['msg'] = u'请先添加专业分包和劳务分包合同'
    except Exception as e:
        response_data['issuc'] = False
        response_data['msg'] = u'添加失败'
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delcalcrelation(request):
    response_data = {}
    calc_id = request.GET.get('calc_id','')
    if calc_id=='':
        response_data['issuc'] = False
        response_data['msg'] = '参数异常'
    else:
        try:
            response_data['issuc'] = True
            CalRelation.objects.filter(id=calc_id).delete()
        except Exception as e:
            # raise e
            response_data['issuc'] = False
            response_data['msg'] = '已关联施工预算，不可删除'
            traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def addbudget_requirerow(request):
    response_data={}
    response_data['issuc'] = True
    require_row_name = request.GET.get('require_row_name','')
    require_row_category = request.GET.get('require_row_category','')
    require_row_bqitem = request.GET.get('require_row_bqitem','')
    require_row_scitem = request.GET.get('require_row_scitem','')
    require_row_valueunit = request.GET.get('require_row_valueunit','')
    require_row_costunit = request.GET.get('require_row_costunit','')
    require_row_desc = request.GET.get('require_row_desc','')
    if CalRelation.objects.filter(bqItem_id=require_row_bqitem).count()==0 and CalRelation.objects.filter(SCItem_id=require_row_scitem).count()==0:
        CalRelation.objects.create(category=require_row_category,name=require_row_name,costUnit=require_row_costunit,valueUnit=require_row_valueunit,differParam=require_row_desc,bqItem_id=require_row_bqitem,SCItem_id=require_row_scitem)
    else:
        response_data['issuc'] = False
        response_data['msg'] = '该项已经存在，请勿重复添加'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def doanalysis_budgetrequirement(docid):
    response_data = {}
    docobj = Document.objects.get(id = docid)
    dicitem = docobj.docdirectory.all()[0]
    path = UPLOAD_DIR+getdirtree(dicitem.id)[1:]+docobj.name
    wb = load_workbook(path)
    sheetname = wb.get_sheet_names()[0]
    sheetContent = wb[sheetname]
    row = sheetContent.rows
    queryList = []
    try:
        category = ''
        currname = ''
        calrelationitem = None
        for index,cell in enumerate(row):
            if index>=1:
                detail = str(cell[0].value)
                detail = detail
                if detail==None or detail=='' or detail=='None':
                    detail = category
                else:
                    category = detail
                name = str(cell[1].value).strip()
                if name==None or name == 'None':
                    name = currname
                else:
                    currname = name
                projects_requires = cell[2].value
                code = cell[4].value
                unit = cell[3].value
                description = cell[2].value
                differParam = cell[5].value

                if description=='分建成本':
                    if code != None:
                        if PactFenbao.objects.filter(name=code.strip()).count()>0:
                            SCItemId = PactFenbao.objects.filter(name=code.strip())[0].id
                            if calrelationitem != None:
                                if CalRelation.objects.filter(bqItem_id=calrelationitem.bqItem.id,SCItem_id=SCItemId).count()<=0:
                                    calrelationitem.SCItem_id = SCItemId
                                    calrelationitem.costUnit = str(cell[3].value)
                                    calrelationitem.toTaskOrder = 0
                                    calrelationitem.save()
                                else:
                                    CalRelation.objects.get(id=calrelationitem.id).delete()
                    else:
                        calrelationitem = None
                        print u'列项为空'
                elif description == '任务单':
                    if code != None:
                        if PactLabour.objects.filter(name=code).count()>0:
                            lpItemid = PactLabour.objects.filter(name=code)[0].id
                            if calrelationitem != None:
                                if CalRelation.objects.filter(bqItem_id=calrelationitem.bqItem.id,lpItem_id=lpItemid).count()<=0:
                                    print u'添加到计算要求'
                                    calrelationitem.lpItem_id = lpItemid
                                    calrelationitem.costUnit = str(cell[3].value)
                                    calrelationitem.toTaskOrder = 1
                                    calrelationitem.save()
                                else:
                                    print u'已存在lpItem_id---删除'
                                    CalRelation.objects.get(id=calrelationitem.id).delete()
                            else:
                                print u'异常'
                    else:
                        print u'劳动分包为空'
                elif description=='产值' or description==None:
                    print u'产值'+str(code)
                    if code == None:
                        if calrelationitem!=None:
                            code = calrelationitem.bqItem.BQItem_Code
                            unit = calrelationitem.valueUnit
                    if code !=None:
                        if BqItem.objects.filter(BQItem_Code=code.strip()).count()>0:
                            bqItemobj = BqItem.objects.filter(BQItem_Code=code.strip())[0]
                            calrelationitem = CalRelation.objects.create(name=name,category=detail,valueUnit=unit,differParam=differParam,bqItem_id=bqItemobj.id)
                        else:
                            calrelationitem = None
                            print u'没有bqItemId关联'
    except Exception as e:
        traceback.print_exc()


def get_budget_tree(request):
    response_data = {}
    response_data['issuc'] = True
    budgetlist = Budget.objects.all()#分包合同
    budgettree = {}
    if len(budgetlist)>0:
        budgettree["id"] = "root"
        budgettree["text"] = '工程量统计'
        budgettree["icon"] = "/img/buildings2.png"
        budgettree["state"] = {'opened': True}
        child_list = []
        for each in budgetlist:
            child_data = {}
            child_data["id"] = each.id
            child_data["text"] = each.name
            child_data["icon"] = "/img/hetong.png"
            child_list.append(child_data)
        huizong_data = {}
        huizong_data["id"] = 'budgethuizong'
        huizong_data["text"] = u'工程量汇总'
        huizong_data["icon"] = "/img/hetong.png"
        huizong_data["state"] = {'opened': True}
        # huizong_data['children'] = child_list
        cost_data = {}
        cost_data["id"] = 'chengben'
        cost_data["text"] = u'Revit导出量'
        cost_data["icon"] = "/img/hetong.png"
        cost_data["state"] = {'opened': True}
        cost_data['children'] = child_list

        rule_data = {}
        rule_data["id"] = 'rule'
        rule_data["text"] = u'工程量对比规则'
        rule_data["icon"] = "/img/hetong.png"
        rule_data["state"] = {'opened': True}

        budgettreechildren = []
        budgettreechildren.append(huizong_data)
        budgettreechildren.append(cost_data)
        budgettreechildren.append(rule_data)
        budgettree['children'] = budgettreechildren
    elif CalRelation.objects.filter().count()>0:
        budgettree["id"] = "root"
        budgettree["text"] = '工程量统计'
        budgettree["icon"] = "/img/buildings2.png"
        budgettree["state"] = {'opened': True}
        child_list = []
        huizong_data = {}
        huizong_data["id"] = 'budgethuizong'
        huizong_data["text"] = u'工程量汇总'
        huizong_data["icon"] = "/img/hetong.png"
        rule_data = {}
        rule_data["id"] = 'rule'
        rule_data["text"] = u'工程量对比规则'
        rule_data["icon"] = "/img/hetong.png"
        rule_data["state"] = {'opened': True}

        child_list.append(huizong_data)
        child_list.append(rule_data)
        budgettree['children'] = child_list
    else:
        budgettree = []
    return HttpResponse(json.dumps(budgettree,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
@csrf_exempt
def getbudgetComponentQuantities(request):
    response_data = {}
    budget_id = int(request.POST.get('param[budget_id]','0'))
    draw = int(request.POST.get('draw','1'))
    start = int(request.POST.get('start','0'))
    pagesize = int(request.POST.get('length','10'))
    end = start+pagesize
    if budget_id==0:
        if Budget.objects.all().count()>0:
            budget_id = Budget.objects.all()[0].id
            datalist = ComponentQuantities.objects.filter(budget_id=budget_id)[start:end]
            le = ComponentQuantities.objects.filter(budget_id=budget_id).count()
            response_data['data'] = serializer(datalist)
            response_data['draw'] = draw
            response_data['recordsTotal'] = le
            response_data['recordsFiltered'] = le
        else:
            response_data['data'] = {}
            response_data['draw'] = 1
            response_data['recordsTotal'] = 0
            response_data['recordsFiltered'] = 0
    else:
        datalist = ComponentQuantities.objects.filter(budget_id=budget_id)[start:end]
        le = ComponentQuantities.objects.filter(budget_id=budget_id).count()
        response_data['data'] = serializer(datalist)
        response_data['draw'] = draw
        response_data['recordsTotal'] = le
        response_data['recordsFiltered'] = le
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''施工预算管理结束'''
def getScitemsBycompany(request):#获取分包单位的分包项
    response_data = {}
    response_data['issuc'] = True
    company_id = int(request.GET.get('company_id','0'))
    pacts = Pact.objects.filter(cooperation_id=company_id,type_id=5)
    sclist = []
    try:
        for pactitem in pacts:
            # print pactitem.name
            scpacts = PactFenbao.objects.filter(Q(pact_id=pactitem.id)&~Q(parent_code=''))
            for scitem in scpacts:
                sclist.append(scitem)
        response_data['scitems'] = serializer(sclist)
    except Exception as e:
        response_data['scitems'] = []
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def getLaborpactsBycompany(request):#获取分包单位的劳务项
    response_data = {}
    response_data['issuc'] = True
    company_id = int(request.GET.get('company_id','0'))
    pacts = Pact.objects.filter(cooperation_id=company_id,type=2)
    laborlist = []
    try:
        for pactitem in pacts:
            loborpacts = PactLabour.objects.filter(pact_id=pactitem.id)
            for laboritem in loborpacts:
                laborlist.append(laboritem)
        response_data['labors'] = serializer(laborlist)
    except Exception as e:
        response_data['labors'] = []
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''任务管理单开始'''
@csrf_exempt
@login_required(login_url="/login/")
def taskmanagerlist(request):
    templateName = 'TaskAndFlow/flowtemplate/business_taskmanager.html'
    scpacts = PactFenbao.objects.filter(~Q(parent_code=''))
    laborpacts = PactLabour.objects.filter()
    reports = Report.objects.filter(~Q(locked=1)).order_by('report_time')
    bqlist = BqItem.objects.filter().order_by('id')
    companys = Company.objects.all()
    professionals = UserMajor.objects.all()
    pacttypes = PactType.objects.all()
    version = time.time()
    return render_to_response(templateName, RequestContext(request, locals()))
def get_task_tree(request):
    try:
        id = request.GET.get('id', '')
        response_data = {}
        child_list = []

        if id == '#':
            response_data["id"] = "maint"
            response_data["text"] = "智慧建造平台"
            response_data["icon"] = "/img/buildings2.png"
            response_data["state"] = {'opened': True}
            treelist = PactSpace.objects.filter().order_by('order_no')

            for each in treelist:
                child_data = {}
                child_data["id"] = 'space-'+str(each.id)
                child_data["text"] = each.name
                child_data["icon"] = "/img/buildings2.png"
                child_data["state"] = {'opened': True}
                child_data_list = []
                children_datazx = {}
                children_datazx["id"] = 'zixing-'+str(each.id)
                children_datazx["text"] = '自行'
                children_datazx["icon"] = "/img/buildings2.png"
                # children_datazx["state"] = {'opened': True}
                children_datazx["children"] = True
                children_datafb = {}
                children_datafb["id"] = 'fenbao-'+str(each.id)
                children_datafb["text"] = '分包'
                children_datafb["icon"] = "/img/buildings2.png"
                # children_datafb["state"] = {'opened': True}
                children_datafb["children"] = True
                child_data_list.append(children_datafb)
                child_data_list.append(children_datazx)
                child_data["children"] = child_data_list
                child_list.append(child_data)

            response_data["children"] = child_list
        else:
            print id
            idarr = id.split('-')
            print idarr
            child_list = []
            if idarr[0] == 'fenbao':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=2):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/buildings2.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)
            elif idarr[0] == 'zixing':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=1):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/buildings2.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)
            elif idarr[0] == 'zixingyg' or idarr[0] == 'fenbaoyg':
                #获取报表
                reports = Report.objects.filter(pact_id=idarr[2])
                for item in reports:
                    print item.start_time.month
                    reportitem_data = {}
                    reportitem_data["id"] = idarr[0]+'report-'+idarr[1]+'-'+idarr[2]+'-'+str(item.id)
                    reportitem_data["text"] = str(item.start_time.month)+u'月成本报表'
                    reportitem_data["icon"] = "/img/buildings2.png"
                    reportitem_data["state"] = {'opened': False}
                    reportitem_data['children'] = True
                    child_list.append(reportitem_data)
                # child_list = tasktreemanager(idarr[2])
            elif idarr[0] == 'zixingygreport' or idarr[0] == 'fenbaoygreport':
                fenjiancb = {}
                laowucb = {}
                fenjiancb["id"] = 'fenjiancb-'+idarr[1]+'-'+idarr[2]+'-'+idarr[3]
                fenjiancb["text"] = u'分建成本'
                fenjiancb["icon"] = "/img/hetong.png"
                fenjiancb["state"] = {'opened': True}
                fenjiancb["children"] = []
                laowucb["id"] = 'laowucb-'+idarr[1]+'-'+idarr[2]+'-'+idarr[3]
                laowucb["text"] = u'劳务成本'
                laowucb["icon"] = "/img/hetong.png"
                laowucb["state"] = {'opened': True}
                laowucb["children"] = []

                shiwul = {}
                noshiwul = {}

                shiwul["id"] = 'shiwuliang-'+idarr[1]+'-'+idarr[2]+'-'+idarr[3]
                shiwul["text"] = u'实物量任务单'
                shiwul["icon"] = "/img/hetong.png"
                shiwul["state"] = {'opened': True}

                noshiwul["id"] = 'noshiwuliang-'+idarr[1]+'-'+idarr[2]+'-'+idarr[3]
                noshiwul["text"] = u'非实物量任务单'
                noshiwul["icon"] = "/img/hetong.png"
                noshiwul["state"] = {'opened': True}

                laowucb["children"].append(shiwul)
                laowucb["children"].append(noshiwul)

                child_list.append(fenjiancb)
                child_list.append(laowucb)
            elif idarr[0] == 'fenbaoyg222':
                child_list = tasktreemanager(idarr[2])

            response_data = child_list

        return HttpResponse(json.dumps(response_data), content_type="application/json")
    except:
        traceback.print_exc()

@csrf_exempt
def gettabletasks(request):
    response_data = {}
    search_company = int(request.POST.get('param[search_company]','0'))
    search_physical = int(request.POST.get('param[search_physical]','1'))
    search_report = int(request.POST.get('param[search_report]','0'))
    search_task = int(request.POST.get('param[search_task]','0'))
    draw = int(request.POST.get('draw','1'))
    start = int(request.POST.get('start','0'))
    pagesize = int(request.POST.get('length','10'))
    end = start+pagesize
    datalist = []
    response_data['data'] = {}
    response_data['draw'] = draw
    response_data['recordsTotal'] = 0
    response_data['recordsFiltered'] = 0
    le = 0
    if search_physical == 1:
        orderlist = []
        shiwulinagfilter = TaskOrderPhysical.objects.filter(report_id=search_report)
        if search_company is not 0:
            shiwulinagfilter = shiwulinagfilter.filter(company_id=search_company)

        for taskitem in shiwulinagfilter:
            item = {}
            item['company'] = taskitem.company.name
            item['issuing_time'] = taskitem.issuing_time
            item['quantities'] = taskitem.quantities
            item['professional'] = taskitem.major.name
            item['price'] = taskitem.pact_price
            item['tasktype'] = 1
            item['reportlocked'] = taskitem.report.locked
            item['reportname'] = taskitem.report.name
            item['name'] = taskitem.sc.name
            item['id'] = taskitem.id
            orderlist.append(item)
        datalist = orderlist[start:end]
        le = len(orderlist)
        response_data['data'] = serializer(datalist)
        response_data['draw'] = draw
        response_data['recordsTotal'] = le
        response_data['recordsFiltered'] = le
        return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
    elif search_physical == 0:
        orderlist = []
        noshiwulinagfilter = TaskOrderNophysical.objects.filter(report_id=search_report)
        print noshiwulinagfilter
        if search_company is not 0:
            noshiwulinagfilter = noshiwulinagfilter.filter(company_id=search_company)

        print noshiwulinagfilter
        for taskitem in noshiwulinagfilter:
            item = {}
            item['company'] = taskitem.company.name
            item['tasktype'] = 2
            item['issuing_time'] = taskitem.issuing_time
            item['quantities'] = taskitem.quantities
            item['professional'] = taskitem.professional.name
            item['price'] = taskitem.workprice
            item['reportlocked'] = taskitem.report.locked
            item['reportname'] = taskitem.report.name
            item['name'] = taskitem.worktype
            item['id'] = taskitem.id
            orderlist.append(item)

        datalist = orderlist[start:end]
        le = len(orderlist)
        response_data['data'] = serializer(datalist)
        response_data['draw'] = draw
        response_data['recordsTotal'] = le
        response_data['recordsFiltered'] = le
        return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")



def getTaskorderByid(request):
    response_data = {}
    response_data['issuc'] = True
    taskorder_id = request.GET.get('id','')
    taskorder_type = int(request.GET.get('tasktype','1'))
    if taskorder_id!='':
        info = {}
        if taskorder_type==1:
            taskorderinfo = TaskOrderPhysical.objects.get(id=taskorder_id)
            info['price'] = float(taskorderinfo.pact_price)
            info['professional_id'] = taskorderinfo.major_id
            info['pact_id'] = taskorderinfo.sc.name
        else:
            taskorderinfo = TaskOrderNophysical.objects.get(id=taskorder_id)
            info['price'] = float(taskorderinfo.workprice)
            info['worktype'] = taskorderinfo.worktype
            info['professional_id'] = taskorderinfo.professional_id
            # info['pact_id'] = taskorderinfo.sc.name
        info['id'] = taskorderinfo.id
        info['issuing_time'] = datetime.datetime.strftime(taskorderinfo.issuing_time,'%Y-%m-%d')
        info['company_id'] = taskorderinfo.company_id
        info['report_id'] = taskorderinfo.report_id
        info['report_name'] = taskorderinfo.report.name


        info['quantities'] = taskorderinfo.quantities
        info['unit'] = taskorderinfo.unit
        info['description'] = taskorderinfo.description
        response_data['taskorder'] = info
    else:
        response_data['issuc'] = False
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def addTaskorder(request):
    response_data = {}
    response_data['issuc'] = True
    operatetype = int(request.GET.get('operatetype',1))
    taskorder_pactprice = request.GET.get('taskorder_pactprice')
    taskorder_issuetime = request.GET.get('taskorder_issuetime')
    taskorder_professional = request.GET.get('taskorder_professional')
    taskorder_company = request.GET.get('taskorder_company')
    taskorder_description = request.GET.get('taskorder_description')
    taskorder_pact = request.GET.get('taskorder_pact')
    taskorder_report = request.GET.get('taskorder_report')
    taskorder_quantities = request.GET.get('taskorder_quantities')
    taskorder_unit = request.GET.get('taskorder_unit')
    taskorderNophysical_worktype = request.GET.get('taskorderNophysical_worktype')

    try:
        reportitem = Report.objects.get(id=taskorder_report)
        if reportitem.locked==1:
            response_data['issuc'] = False
            response_data['msg'] = u'关联产值已经上报了'
        else:
            if operatetype==1:
                taskordermodel = TaskOrderPhysical.objects.create(unit=taskorder_unit,sc_id=taskorder_pact,report_id=taskorder_report,quantities=taskorder_quantities,description=taskorder_description,pact_price=taskorder_pactprice,company_id=taskorder_company,major_id=taskorder_professional,issuing_time=taskorder_issuetime)
            else:
                TaskOrderNophysical.objects.create(worktype=taskorderNophysical_worktype,unit=taskorder_unit,report_id=taskorder_report,quantities=taskorder_quantities,description=taskorder_description,workprice=taskorder_pactprice,company_id=taskorder_company,professional_id=taskorder_professional,issuing_time=taskorder_issuetime)
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def updateTaskorder(request):
    response_data = {}
    response_data['issuc'] = True
    taskorder_type = int(request.GET.get('taskorder_type','1'))
    taskorder_id = request.GET.get('taskorder_id')
    taskorder_price = request.GET.get('taskorder_price')
    taskorder_issuetime = request.GET.get('taskorder_issuetime')
    taskorder_professional = request.GET.get('taskorder_professional')
    taskorder_company = request.GET.get('taskorder_company')
    taskorder_description = request.GET.get('taskorder_description')
    taskorder_unit = request.GET.get('taskorder_unit')
    taskorder_report = request.GET.get('taskorder_report')
    taskorder_pact = request.GET.get('taskorder_pact')
    taskorder_quantities = request.GET.get('taskorder_quantities')
    taskorder_worktype = request.GET.get('taskorder_worktype')
    try:
        if taskorder_type==1:
            taskordermodel = TaskOrderPhysical.objects.get(id=taskorder_id)
            taskordermodel.pact_price = taskorder_price
            # taskordermodel.company_id = taskorder_company
            taskordermodel.description = taskorder_description
            taskordermodel.major_id = taskorder_professional
            taskordermodel.issuing_time = taskorder_issuetime
            # taskordermodel.sc_id = taskorder_pact
            # taskordermodel.report_id = taskorder_report
            taskordermodel.unit = taskorder_unit
            taskordermodel.quantities = taskorder_quantities
            taskordermodel.save()
        else:
            taskordermodel = TaskOrderNophysical.objects.get(id=taskorder_id)
            taskordermodel.workprice = taskorder_price
            # taskordermodel.company_id = taskorder_company
            taskordermodel.description = taskorder_description
            taskordermodel.professional_id = taskorder_professional
            taskordermodel.issuing_time = taskorder_issuetime
            # taskordermodel.sc_id = taskorder_pact
            # taskordermodel.report_id = taskorder_report
            taskordermodel.unit = taskorder_unit
            taskordermodel.quantities = taskorder_quantities
            taskordermodel.worktype = taskorder_worktype
            taskordermodel.save()
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delTaskorderByid(request):
    response_data = {}
    response_data['issuc'] = True
    taskorder_id = request.GET.get('id','')
    tasktype = int(request.GET.get('tasktype','1'))
    try:
        if taskorder_id!='':
            if tasktype==1:
                TaskOrderPhysical.objects.filter(id=taskorder_id).delete()
            else:
                TaskOrderNophysical.objects.get(id=taskorder_id).delete()
        else:
            response_data['issuc'] = False
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''任务管理单结束'''
'''施工产值计报开始'''
@login_required(login_url="/login/")
def chanzhilist(request):
    templateName = 'TaskAndFlow/flowtemplate/business_chanzhilist.html'
    version = time.time()
    if Report.objects.filter().count()>0:
        report_id = Report.objects.filter()[0].id
    else:
        report_id = 'null'
    bqlist = BqItem.objects.filter()
    if bqlist:
        bqitem = BqItem.objects.filter()[0]
    scpacts = PactFenbao.objects.filter(~Q(parent_code=''))
    resourcelist = RtItemResource.objects.filter()
    budgetlist = Pact.objects.filter(type_id=1)
    return render_to_response(templateName, RequestContext(request, locals()))
def getqingdancodeinfo(request):
    response_data = {}
    bqid = request.GET.get('id','')
    response_data['issuc'] = True
    if bqid == '':
        response_data['issuc'] = False
        response_data['issuc'] = u'参数错误'
    else:
        item = BqItem.objects.get(id=bqid)
        response_data['bqitem'] = serializer(item)
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def getSciteminfo(request):
    response_data = {}
    bqid = request.GET.get('id','')
    response_data['issuc'] = True
    if bqid == '':
        response_data['issuc'] = False
        response_data['issuc'] = u'参数错误'
    else:
        item = PactFenbao.objects.get(id=bqid)
        response_data['scitem'] = serializer(item)
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delReport(request):
    response_data = {}
    response_data['issuc'] = True
    report_id = int(request.GET.get('report_id','0'))
    try:
        if report_id==0 or Report.objects.filter(id=report_id).count()<=0:
            response_data['issuc'] = False
            response_data['msg'] = '参数异常或者报告不存在'
        else:
            reportmodel = Report.objects.filter(id=report_id)[0]
            # print (reportmodel.locked == 1)
            if reportmodel.locked == 1:
                response_data['issuc'] = False
                response_data['msg'] = '该报告已经上报不能删除'
            else:
                for reportbqitem in ReportBqitem.objects.filter(report=report_id):
                    ReportRtitem.objects.filter(reportbq_id=reportbqitem.id).delete()
                ReportBqitem.objects.filter(report_id=report_id).delete()
                ReportResItem.objects.filter(report_id=report_id).delete()
                ReportCostItem.objects.filter(report_id=report_id).delete()

                CostSeparate.objects.filter(report_id=report_id).delete()
                TaskOrderNophysical.objects.filter(report_id=report_id).delete()
                TaskOrderPhysical.objects.filter(report_id=report_id).delete()
                Report.objects.filter(id=report_id).delete()
    except Exception as e:
        response_data['issuc'] = False
        response_data['msg'] = '程序异常'
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delReportBqitem(request):
    response_data = {}
    response_data['issuc'] = True
    deltype = int(request.GET.get('deltype','0'))
    report_bqid = int(request.GET.get('report_bqid','0'))
    report_rtid = int(request.GET.get('report_rtid','0'))
    if deltype==0:
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        if deltype==1:#删除清单
            if report_bqid!=0:
                ReportRtitem.objects.filter(reportbq_id=report_bqid).delete()
                ReportBqitem.objects.filter(id=report_bqid).delete()
        elif deltype==2:
            ReportRtitem.objects.filter(id=report_rtid).delete()
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'错误的操作'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
@csrf_exempt
def updateReportBqitem(request):
    response_data = {}
    id = request.POST.get('pk','')
    name = request.POST.get('name')
    value = request.POST.get('value','')
    operate = name.split('-')
    if operate[0] == 'price':#修改清单价格
        # print u'修改清单价格'
        if operate[1] == '1':
            print u'修改清单价格'
            bqitem = ReportBqitem.objects.get(id=id)
            bqitem.unitprice = value
            bqitem.save()
        elif operate[1] == '2':
            print u'修改定额价格'
            rtitem = ReportRtitem.objects.get(id=id)
            rtitem.unitprice = value
            rtitem.save()
    elif operate[0] == 'nums':
        print u'修改清单工程量'
        if operate[1] == '1':
            bqitem = ReportBqitem.objects.get(id=id)
            try:
                PostUrl = 'http://121.41.22.17:17531/api/ValueReportChanges'
                cookie = cookielib.CookieJar()
                handler = urllib2.HTTPCookieProcessor(cookie)
                opener = urllib2.build_opener(handler)
                postData={
                    'reportbqcode':bqitem.bq.BQItem_Code,
                    'reportid':bqitem.report.id,
                    'bqamount':value,
                    'schema':str(DATABASES['pms']['NAME'])
                }
                print postData
                headers = {}
                data = urllib.urlencode(postData)
                request = urllib2.Request(PostUrl,data)
                response = opener.open(request)
                result = response.read()
                print result
            except Exception as e:
                traceback.print_exc()
        elif operate[1] == '2':
            print u'修改定额工程量'
            rtitem = ReportRtitem.objects.get(id=id)
            rtitem.quantity = value
            rtitem.save()
    response_data['issucc'] = 'true'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''添加报表清单'''
def addReportBqitem(request):
    response_data = {}
    response_data['issuc'] = True
    bqid = request.GET.get('bqid','')
    name = request.GET.get('name')
    unit = request.GET.get('unit')
    amount = request.GET.get('amount')
    price = request.GET.get('price')
    report_id = request.GET.get('report_id','')
    if report_id == '' or bqid == '':
        response_data['issuc'] = False
        response_data['msg'] = u'参数不正确'
    else:
        if ReportBqitem.objects.filter(report_id=report_id,bq_id=bqid).count()>0:
            response_data['issuc'] = False
            response_data['msg'] = u'该清单已存在'
        else:
            reportitem = Report.objects.get(id=int(report_id))
            if reportitem.locked==0 or reportitem.locked==None or reportitem.locked=='':
                bqitemmodel = BqItem.objects.get(id=bqid)
                rbqm = ReportBqitem.objects.create(code=bqitemmodel.BQItem_Code,name=name,unit=unit,quantity=amount,unitprice=price,report_id=report_id,bq_id=bqid)
                if RtItem.objects.filter(BQ_Item_id=bqitemmodel.id)>0:
                    rqs = RtItem.objects.filter(BQ_Item_id=bqitemmodel.id)
                    execute_list = []
                    for rtitem in rqs:
                        execute_list.append(ReportRtitem(code=rtitem.RQItem_Code,name=rtitem.RT_Item_Name,unit=rtitem.unit,quantity=rtitem.quantities,unitprice=rtitem.unitprice,rtitem_id=rtitem.id,reportbq_id=rbqm.id))
                    ReportRtitem.objects.bulk_create(execute_list)
            else:
                response_data['issuc'] = False
                response_data['msg'] = u'该月产值已经上报，不能修改'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def lockReport(request):
    response_data = {}
    response_data['issuc'] = True
    report_id = int(request.GET.get('report_id','0'))
    if report_id==0:
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        reportitem = Report.objects.get(id=report_id)
        if reportitem.locked == 0 or reportitem.locked==None or reportitem.locked=='':
            reportitem.locked = 1
            reportitem.save()
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'该产值报告已经上报过了'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def addReportRes(request):
    response_data = {}
    response_data['issuc'] = True
    resourceid = request.GET.get('resourceid','')
    name = request.GET.get('name')
    amount = request.GET.get('amount')
    price = request.GET.get('price')
    report_id = request.GET.get('report_id','')
    if report_id == '' or resourceid == '':
        response_data['issuc'] = False
        response_data['msg'] = u'参数不正确'
    else:
        if RtItemResource.objects.filter(id=resourceid).count()>0:
            reportitem = Report.objects.get(id=int(report_id))
            if reportitem.locked == 0 or reportitem.locked==None or reportitem.locked=='':
                rtresitem = RtItemResource.objects.filter(id=resourceid)[0]
                ReportResItem.objects.create(code=rtresitem.Code,name=name,unit=rtresitem.unit,amount=amount,unitprice=price,resource_id=resourceid,report_id=report_id)
            else:
                response_data['issuc'] = False
                response_data['msg'] = u'该月产值已经上报，不能修改'
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'无效的资源编号'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
@csrf_exempt
def updateReportResitem(request):
    response_data = {}
    id = request.POST.get('pk','')
    operate = request.POST.get('name')
    value = request.POST.get('value','')
    resitem = ReportResItem.objects.get(id=id)
    if operate == 'price':#修改清单价格
        resitem.unitprice = value
    elif operate == 'nums':
        resitem.amount = value
    resitem.save()
    response_data['issucc'] = 'true'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delReportRes(request):
    response_data = {}
    response_data['issuc'] = True
    report_sourceid = int(request.GET.get('report_sourceid','0'))
    if report_sourceid==0:
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        ReportResItem.objects.filter(id=report_sourceid).delete()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

def addReportCost(request):
    response_data = {}
    response_data['issuc'] = True
    name = request.GET.get('name')
    money = request.GET.get('money')
    rate = request.GET.get('rate')
    ratedescription = request.GET.get('ratedescription')
    scid = request.GET.get('scid')
    report_id = request.GET.get('report_id','')
    if report_id == '' or name == '':
        response_data['issuc'] = False
        response_data['msg'] = u'参数不正确'
    else:
        reportitem = Report.objects.get(id=int(report_id))
        if reportitem.locked == 0 or reportitem.locked==None or reportitem.locked=='':
            if ReportCostItem.objects.filter(sc_id=scid,report_id=report_id).count()<=0:
                scitem = PactFenbao.objects.get(id=scid)
                ReportCostItem.objects.create(money=money,rate=rate,ratedescription=ratedescription,name=name,report_id=report_id,sc_id=scid,quanlity=scitem.designBqs,unitprice=scitem.price,unit=scitem.unit)
            else:
                response_data['issuc'] = False
                response_data['msg'] = u'该月产值报告中已存在该费用项'
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'该月产值已经上报，不能修改'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
'''添加分建成本'''
def addSeparateCost(request):
    response_data = {}
    response_data['issuc'] = True
    unit = request.GET.get('unit')
    price = request.GET.get('price')
    money = request.GET.get('money')
    quantities = request.GET.get('quantities')
    scid = request.GET.get('scid')
    companyid = request.GET.get('companyid')
    report_id = request.GET.get('report_id','')
    if report_id == '' or companyid == '':
        response_data['issuc'] = False
        response_data['msg'] = u'参数不正确'
    else:
        reportitem = Report.objects.get(id=int(report_id))
        if reportitem.locked == 0 or reportitem.locked==None or reportitem.locked=='':
            if CostSeparate.objects.filter(sc_id=scid,report_id=report_id).count()<=0:
                scitem = PactFenbao.objects.get(id=scid)
                CostSeparate.objects.create(pro_name=scitem.name,report_id=report_id,sc_id=scid,quantities=quantities,price=price,unit=unit,money=money,company_id=companyid)
            else:
                response_data['issuc'] = False
                response_data['msg'] = u'该月产值报告中已存在该分建成本'
        else:
            response_data['issuc'] = False
            response_data['msg'] = u'该月产值已经上报，无法添加分建成本'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

def delReportCost(request):
    response_data = {}
    response_data['issuc'] = True
    report_costid = int(request.GET.get('report_costid','0'))
    if report_costid==0:
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        ReportCostItem.objects.filter(id=report_costid).delete()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def delSeparateCost(request):
    response_data = {}
    response_data['issuc'] = True
    costid = int(request.GET.get('id','0'))
    if costid==0:
        response_data['issuc'] = False
        response_data['msg'] = u'参数异常'
    else:
        CostSeparate.objects.filter(id=costid).delete()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

def calculateReport(request):
    response_data = {}
    response_data['issuc'] = True
    report_name = request.GET.get('report_name','')
    start_time = request.GET.get('start_time','')
    end_time = request.GET.get('end_time','')
    budget_id = request.GET.get('budget_id','')
    if report_name=='' or start_time=='' or end_time=='':
        response_data['issuc'] = False
    try:
        PostUrl = 'http://121.41.22.17:17531/api/ValueReport'
        cookie = cookielib.CookieJar()
        handler = urllib2.HTTPCookieProcessor(cookie)
        opener = urllib2.build_opener(handler)
        postData={
            'report_name':report_name,
            'start_time':start_time,
            'end_time':end_time,
            'pactId':int(budget_id),
            'schema':str(DATABASES['pms']['NAME'])
        }
        headers = {}
        data = urllib.urlencode(postData)
        request = urllib2.Request(PostUrl,data)
        response = opener.open(request)
        result = response.read()

    except Exception as e:
        traceback.print_exc()
        print u'计算产值调用接口异常'
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def reporttreemanager(reportid,type):
    each = Report.objects.get(id=reportid)#分包合同
    child_list = []
    print 1111,reportid
    if type=='chanzhi':
        bqs_data = {}#清单
        bqs_data["id"] = 'bqs-'+str(each.id)
        bqs_data["text"] = '施工项目清单'
        bqs_data["icon"] = "/img/hetong.png"

        res_data = {}#资源
        res_data["id"] = 'res-'+str(each.id)
        res_data["text"] = '工料机汇总表'
        res_data["icon"] = "/img/hetong.png"

        res2_data = {}#定额工料机
        res2_data["id"] = 'thound2-'+str(each.id)
        res2_data["text"] = '2000定额工料机'
        res2_data["icon"] = "/img/hetong.png"

        rate_data = {}#费率
        rate_data["id"] = 'rate-'+str(each.id)
        rate_data["text"] = '费率表'
        rate_data["icon"] = "/img/hetong.png"

        child_list.append(bqs_data)
        child_list.append(rate_data)
        child_list.append(res_data)
        child_list.append(res2_data)
    else:
        company_list = []
        res = []
        physicallist = []
        nophysicallist = []
        costsperatelist = []
        if CostSeparate.objects.filter(report_id=each.id).count()>0:
            costsperatelist = CostSeparate.objects.filter(report_id=each.id)
            for item in costsperatelist:
                company_data = {}#公司
                company_data["id"] = 'company-'+str(each.id)+'-'+str(item.company.id)
                company_data["text"] = item.company.name
                company_data["icon"] = "/img/hetong.png"
                company_data["state"] = {'opened': True}

                fenjian_data = {}#分建成本
                fenjian_data["id"] = 'fenjian-'+str(each.id)+'-'+str(item.company.id)
                fenjian_data["text"] = '分建成本'
                fenjian_data["icon"] = "/img/hetong.png"

                laowu_data = {}#劳务成本
                laowu_data["id"] = 'laowu-'+str(each.id)+'-'+str(item.company.id)
                laowu_data["text"] = '劳务成本'
                laowu_data["icon"] = "/img/hetong.png"

                conpany_children_list = []
                conpany_children_list.append(fenjian_data)
                conpany_children_list.append(laowu_data)
                company_data['children'] = conpany_children_list
                company_list.append(company_data)
        if TaskOrderPhysical.objects.filter(report_id=each.id).count()>0:
            physicallist = TaskOrderPhysical.objects.filter(report_id=each.id)
            for item in physicallist:
                company_data = {}#公司
                company_data["id"] = 'company-'+str(each.id)+'-'+str(item.company.id)
                company_data["text"] = item.company.name
                company_data["icon"] = "/img/hetong.png"
                company_data["state"] = {'opened': True}

                fenjian_data = {}#分建成本
                fenjian_data["id"] = 'fenjian-'+str(each.id)+'-'+str(item.company.id)
                fenjian_data["text"] = '分建成本'
                fenjian_data["icon"] = "/img/hetong.png"

                laowu_data = {}#劳务成本
                laowu_data["id"] = 'laowu-'+str(each.id)+'-'+str(item.company.id)
                laowu_data["text"] = '劳务成本'
                laowu_data["icon"] = "/img/hetong.png"

                conpany_children_list = []
                conpany_children_list.append(fenjian_data)
                conpany_children_list.append(laowu_data)
                company_data['children'] = conpany_children_list
                company_list.append(company_data)
        if TaskOrderNophysical.objects.filter(report_id=each.id).count()>0:
            nophysicallist = TaskOrderNophysical.objects.filter(report_id=each.id)
            for item in nophysicallist:
                company_data = {}#公司
                company_data["id"] = 'company-'+str(each.id)+'-'+str(item.company.id)
                company_data["text"] = item.company.name
                company_data["icon"] = "/img/hetong.png"
                company_data["state"] = {'opened': True}

                fenjian_data = {}#分建成本
                fenjian_data["id"] = 'fenjian-'+str(each.id)+'-'+str(item.company.id)
                fenjian_data["text"] = '分建成本'
                fenjian_data["icon"] = "/img/hetong.png"

                laowu_data = {}#劳务成本
                laowu_data["id"] = 'laowu-'+str(each.id)+'-'+str(item.company.id)
                laowu_data["text"] = '劳务成本'
                laowu_data["icon"] = "/img/hetong.png"

                conpany_children_list = []
                conpany_children_list.append(fenjian_data)
                conpany_children_list.append(laowu_data)
                company_data['children'] = conpany_children_list
                company_list.append(company_data)
        [child_list.append(x) for x in company_list if x not in child_list]#去重
    return child_list

'''产值报表-左侧树，异步加载'''
def get_hazard_tree(request):
    try:
        id = request.GET.get('id', '')
        child_list = []
        if id == '#':
            treelist = PactSpace.objects.filter().order_by('order_no')
            for each in treelist:
                child_data = {}
                child_data["id"] = 'space-'+str(each.id)
                child_data["text"] = each.name
                child_data["icon"] = "/img/buildings2.png"
                child_data["state"] = {'opened': True}
                child_data_list = []
                children_datazx = {}
                children_datazx["id"] = 'zixing-'+str(each.id)
                children_datazx["text"] = '自行'
                children_datazx["icon"] = "/img/buildings2.png"

                children_datazx["children"] = True
                children_datafb = {}
                children_datafb["id"] = 'fenbao-'+str(each.id)
                children_datafb["text"] = '分包'
                children_datafb["icon"] = "/img/buildings2.png"

                children_datafb["children"] = True
                child_data_list.append(children_datafb)
                child_data_list.append(children_datazx)
                child_data["children"] = child_data_list
                child_list.append(child_data)
        else:
            idarr = id.split('-')
            child_list = []
            #fenbao、zixing获取分包、自行中的预算，zixingyg，fenbaoyg获取预算中的产值报告
            if idarr[0] == 'fenbao':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=2):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/buildings2.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)
            elif idarr[0] == 'zixing':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=1):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/buildings2.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)
            elif idarr[0] == 'zixingyg':
                for reportitem in Report.objects.filter(pact_id=idarr[2]).order_by('-start_time'):
                    tree_data = {}
                    tree_data["id"] = 'chanzhi-'+idarr[1]+'-'+idarr[2]+'-'+str(reportitem.id)
                    tree_data["text"] = reportitem.name
                    tree_data["icon"] = "/img/hetong.png"
                    tree_data["state"] = {'opened': False}
                    tree_data['children'] = True

                    child_chengben_data = {}#成本
                    child_chengben_data["id"] = 'chengben-'+idarr[1]+'-'+idarr[2]+'-'+str(reportitem.id)
                    child_chengben_data["text"] = (reportitem.name).replace(u'产值',u'分包成本')
                    child_chengben_data["icon"] = "/img/hetong.png"
                    child_chengben_data["state"] = {'opened': False}
                    child_chengben_data['children'] = True

                    child_list.append(tree_data)
                    # child_list.append(child_chengben_data)
            elif idarr[0] == 'fenbaoyg':
                for reportitem in Report.objects.filter(pact_id=idarr[2]).order_by('-start_time'):
                    tree_data = {}
                    tree_data["id"] = 'chanzhi-'+idarr[1]+'-'+idarr[2]+'-'+str(reportitem.id)
                    tree_data["text"] = reportitem.name
                    tree_data["icon"] = "/img/hetong.png"
                    tree_data["state"] = {'opened': False}
                    tree_data['children'] = True

                    child_chengben_data = {}#成本
                    child_chengben_data["id"] = 'chengben-'+idarr[1]+'-'+idarr[2]+'-'+str(reportitem.id)
                    child_chengben_data["text"] = (reportitem.name).replace(u'产值',u'分包成本')
                    child_chengben_data["icon"] = "/img/hetong.png"
                    child_chengben_data["state"] = {'opened': False}
                    child_chengben_data['children'] = True

                    child_list.append(tree_data)
                    # child_list.append(child_chengben_data)
            elif idarr[0] == 'chanzhi':
                child_list = reporttreemanager(idarr[3],'chanzhi')
            elif idarr[0] == 'chengben':
                child_list = reporttreemanager(idarr[3],'chengben')

        return HttpResponse(json.dumps(child_list), content_type="application/json")
    except:
        traceback.print_exc()

def get_report_tree(request):
    response_data = {}
    tree_list = []
    spacelist = PactSpace.objects.filter().order_by('order_no')
    pro_name = Project.objects.get(id=CURRENT_PROJECT_ID).name
    reporttree_main = {}
    reporttree_main["id"] = "root"
    reporttree_main["text"] = pro_name
    reporttree_main["icon"] = "/img/buildings2.png"
    reporttree_main["state"] = {'opened': True}
    for each in spacelist:
        item = {}
        item["id"] = each.id
        item["text"] = each.name
        item["icon"] = "/img/buildings2.png"
        item["state"] = {'opened': False}
        tree_list.append(item)
        tree_child_data = []
        children_data1 = {}
        children_data1["id"] = 'zixing-'+str(each.id)
        children_data1["text"] = '自行'
        children_data1["icon"] = "/img/buildings2.png"
        children_data1["state"] = {'opened': False}
        if Pact.objects.filter(space_id=each.id,is_self=1).count()>0:
            pactitemlist = []
            for pactitem in Pact.objects.filter(space_id=each.id,is_self=1):
                pactitem_data = {}
                pactitem_data["id"] = 'zixingyg-'+str(each.id)+'-'+str(pactitem.id)
                pactitem_data["text"] = pactitem.name
                pactitem_data["icon"] = "/img/hetong.png"
                pactitem_data["state"] = {'opened': False}
                pactitem_data['children'] = reporttreemanager(pactitem.id)
                pactitemlist.append(pactitem_data)
            children_data1['children'] = pactitemlist

        children_data2 = {}
        children_data2["id"] = 'fenbao-'+str(each.id)
        children_data2["text"] = '分包'
        children_data2["icon"] = "/img/buildings2.png"
        children_data2["state"] = {'opened': False}
        if Pact.objects.filter(space_id=each.id,is_self=2).count()>0:
            pactitemlist = []
            for pactitem in Pact.objects.filter(space_id=each.id,is_self=2):
                pactitem_data = {}
                pactitem_data["id"] = 'fenbaoyg-'+str(each.id)+'-'+str(pactitem.id)
                pactitem_data["text"] = pactitem.name
                pactitem_data["icon"] = "/img/hetong.png"
                pactitem_data["state"] = {'opened': False}
                pactitem_data['children'] = reporttreemanager(pactitem.id)
                pactitemlist.append(pactitem_data)
            children_data2['children'] = pactitemlist
        tree_child_data.append(children_data1)
        tree_child_data.append(children_data2)
        item["children"] = tree_child_data
    reporttree_main['children'] = tree_list
    return HttpResponse(json.dumps(reporttree_main,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
def get_report_account_tree(request):
    response_data = {}
    tree_list = []
    spacelist = PactSpace.objects.filter().order_by('order_no')
    pro_name = Project.objects.get(id=CURRENT_PROJECT_ID).name
    reporttree = {}
    reporttree["id"] = "root"
    reporttree["text"] = pro_name
    reporttree["icon"] = "/img/buildings2.png"
    reporttree["state"] = {'opened': True}
    for each in spacelist:
        item = {}
        item["id"] = each.id
        item["text"] = each.name
        item["icon"] = "/img/buildings2.png"
        item["state"] = {'opened': False}
        tree_list.append(item)
        tree_child_data = []
        children_data1 = {}
        children_data1["id"] = 'zixing-'+str(each.id)
        children_data1["text"] = '自行'
        children_data1["icon"] = "/img/buildings2.png"
        children_data1["state"] = {'opened': False}
        if Pact.objects.filter(space_id=each.id,is_self=1).count()>0:
            pactitemlist = []
            for pactitem in Pact.objects.filter(space_id=each.id,is_self=1):
                pactitem_data = {}
                pactitem_data["id"] = 'zixingyg-'+str(each.id)+'-'+str(pactitem.id)
                pactitem_data["text"] = pactitem.name
                pactitem_data["icon"] = "/img/hetong.png"
                pactitem_data["state"] = {'opened': False}
                pactitemlist.append(pactitem_data)
            children_data1['children'] = pactitemlist

        children_data2 = {}
        children_data2["id"] = 'fenbao-'+str(each.id)
        children_data2["text"] = '分包'
        children_data2["icon"] = "/img/buildings2.png"
        children_data2["state"] = {'opened': False}
        if Pact.objects.filter(space_id=each.id,is_self=2).count()>0:
            pactitemlist = []
            for pactitem in Pact.objects.filter(space_id=each.id,is_self=2):
                pactitem_data = {}
                pactitem_data["id"] = 'fenbaoyg-'+str(each.id)+'-'+str(pactitem.id)
                pactitem_data["text"] = pactitem.name
                pactitem_data["icon"] = "/img/hetong.png"
                pactitem_data["state"] = {'opened': False}
                pactitemlist.append(pactitem_data)
            children_data2['children'] = pactitemlist
        tree_child_data.append(children_data1)
        tree_child_data.append(children_data2)
        item["children"] = tree_child_data
    reporttree['children'] = tree_list
    return HttpResponse(json.dumps(reporttree,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

'''产值报表表格数据获取（用于table分页组件）param[report_type]-获取表格类型 1-清单，2-人材机，3-费率'''
@csrf_exempt
def getbqtablelist(request):
    response_data = {}
    try:
        report_id = int(request.POST.get('param[report_id]','0'))
        report_type = int(request.POST.get('param[report_type]','1'))
        draw = int(request.POST.get('draw','1'))
        start = int(request.POST.get('start','0'))
        pagesize = int(request.POST.get('length','10'))
        end = start+pagesize
        datalist = []
        num = 0
        if report_id != 0:
            reportitem = Report.objects.get(id=report_id)
            if report_type == 1:#获取清单信息
                reportlist = ReportBqitem.objects.filter(report_id=report_id)
                itlist = []
                for reportbqitem in reportlist:
                    new_bqitem = {}
                    new_bqitem = serializer(reportbqitem)
                    new_bqitem['xuhao'] = reportbqitem.bq.no
                    if reportbqitem.isLocked==1 or reportitem.locked==1:
                        new_bqitem['locked'] = 1
                    else:
                        new_bqitem['locked'] = 0
                    new_bqitem['datatype'] = 1
                    itlist.append(new_bqitem)
                    num = num+1
                    if ReportRtitem.objects.filter(reportbq_id=reportbqitem.id).count()>0:
                        rtlists = ReportRtitem.objects.filter(reportbq_id=reportbqitem.id)
                        for reportrtitem in rtlists:
                            new_rtitem = {}
                            new_rtitem = serializer(reportrtitem)
                            new_rtitem['xuhao'] = None
                            new_rtitem['locked'] = reportrtitem.reportbq.isLocked
                            if reportrtitem.reportbq.isLocked==1 or reportitem.locked==1:
                                new_bqitem['locked'] = 1
                            else:
                                new_bqitem['locked'] = 0
                            new_rtitem['quantity'] = reportbqitem.quantity*reportrtitem.rtitem.quantities/reportbqitem.bq.designBqs
                            new_rtitem['datatype'] = 2
                            itlist.append(new_rtitem)
                            num = num+1
                datalist = itlist[start:end]
                le = num
            elif report_type == 2:#获取资源信息
                datalistpre = ReportResItem.objects.filter(report_id=report_id)[start:end]
                rris = ReportResItem.objects.filter(report_id=report_id).values('code').annotate(amount_sum=Sum('amount'))

                for item in rris:
                    itemobj = serializer(ReportResItem.objects.filter(code=item['code'])[0])
                    itemobj['amount'] = item['amount_sum']
                    datalist.append(itemobj)
                print len(rris)
                le = len(rris)
            elif report_type == 3:#获取费用信息
                datalistpre = ReportCostItem.objects.filter(report_id=report_id)[start:end]
                for item in datalistpre:
                    itemobj = serializer(item)
                    itemobj['report_id'] = item.report_id
                    datalist.append(itemobj)
                le = ReportCostItem.objects.filter(report_id=report_id).count()
            elif report_type == 4:#获取分建成本
                company = int(request.POST.get('param[company]','0'))
                datalistpre = CostSeparate.objects.filter(report_id=report_id)
                if company is not 0:
                    datalistpre = datalistpre.filter(company_id=company)
                # datalistpre = ReportCostItem.objects.filter(report_id=report_id)[start:end]
                itemlist = []
                for item in datalistpre:
                    itemobj = {}
                    itemobj['report_id'] = item.report_id
                    itemobj['id'] = item.id
                    itemobj['name'] = item.pro_name
                    itemobj['unit'] = item.unit
                    itemobj['quanlity'] = Decimal(float(item.quantities)).quantize(Decimal('0.00'))
                    itemobj['unitprice'] = Decimal(float(item.price)).quantize(Decimal('0.00'))
                    itemobj['locked'] = reportitem.locked

                    if CalRelation.objects.filter(SCItem_id=item.sc.id).count()>0:
                        calitem = CalRelation.objects.filter(SCItem_id=item.sc.id)[0]
                        itemobj['xuhao'] = calitem.bqItem.no
                        itemobj['code'] = calitem.bqItem.BQItem_Code
                        itemlist.append(itemobj)
                    else:
                        print item.pro_name,item.sc.id
                        itemobj['xuhao'] = None
                        itemobj['code'] = None
                datalist = itemlist[start:end]
                le = len(itemlist)
            elif report_type == 5:#获取劳务成本
                company = int(request.POST.get('param[company]','1'))
                laowulist = []
                if TaskOrderPhysical.objects.filter(report_id=report_id,company_id=company).count()>0:
                    for physicalitem in TaskOrderPhysical.objects.filter(report_id=report_id,company_id=company):
                        itemobj = {}
                        itemobj['report_id'] = report_id
                        itemobj['id'] = physicalitem.id
                        itemobj['tasktype'] = 1
                        itemobj['name'] = physicalitem.sc.name
                        itemobj['unit'] = physicalitem.unit
                        itemobj['quanlity'] = physicalitem.quantities
                        itemobj['unitprice'] = physicalitem.pact_price
                        itemobj['locked'] = physicalitem.report.locked
                        itemobj['xuhao'] = None
                        itemobj['code'] = None
                        laowulist.append(itemobj)
                if TaskOrderNophysical.objects.filter(report_id=report_id,company_id=company).count()>0:
                    for physicalitem in TaskOrderNophysical.objects.filter(report_id=report_id,company_id=company):
                        itemobj = {}
                        itemobj['report_id'] = report_id
                        itemobj['id'] = physicalitem.id
                        itemobj['tasktype'] = 2
                        itemobj['name'] = physicalitem.worktype
                        itemobj['unit'] = physicalitem.unit
                        itemobj['quanlity'] = physicalitem.quantities
                        itemobj['unitprice'] = physicalitem.workprice
                        itemobj['locked'] = physicalitem.report.locked
                        itemobj['xuhao'] = None
                        itemobj['code'] = None
                        laowulist.append(itemobj)
                datalist = laowulist[start:end]
                le = len(datalist)
            elif report_type == 6:#获取2000定额工料机
                datalistpre = ReportResThoundItem.objects.filter(report_id=report_id)[start:end]
                rris = ReportResThoundItem.objects.filter(report_id=report_id).values('code').annotate(amount_sum=Sum('amount'))
                for item in rris:
                    itemobj = serializer(ReportResThoundItem.objects.filter(code=item['code'])[0])
                    itemobj['amount'] = item['amount_sum']
                    datalist.append(itemobj)
                le = 0
            elif report_type == 7:#获取费率
                print u'获取费率'
                hascount = ReportRate.objects.filter(report_id=reportitem.id).count()
                le = hascount
                datalist = ReportRate.objects.filter(report_id=reportitem.id).order_by('num')

        else:#默认获取清单信息
            if Report.objects.filter().count()>0:
                reportitem = Report.objects.filter().order_by("-report_time")[0]
                if report_type==1:
                    reportlist = ReportBqitem.objects.filter(report_id=reportitem.id)
                    itlist = []
                    for reportbqitem in reportlist:
                        new_bqitem = {}
                        new_bqitem = serializer(reportbqitem)
                        new_bqitem['xuhao'] = reportbqitem.bq.no
                        if reportbqitem.isLocked==1 or reportitem.locked==1:
                            new_bqitem['locked'] = 1
                        else:
                            new_bqitem['locked'] = 0
                        new_bqitem['datatype'] = 1
                        itlist.append(new_bqitem)
                        num = num+1
                        if ReportRtitem.objects.filter(reportbq_id=reportbqitem.id).count()>0:
                            rtlists = ReportRtitem.objects.filter(reportbq_id=reportbqitem.id)
                            for reportrtitem in rtlists:
                                new_rtitem = {}
                                new_rtitem = serializer(reportrtitem)

                                new_rtitem['quantity'] = reportbqitem.quantity*reportrtitem.rtitem.quantities/reportbqitem.bq.designBqs
                                new_rtitem['xuhao'] = None
                                if reportrtitem.reportbq.isLocked==1 or reportitem.locked==1:
                                    new_bqitem['locked'] = 1
                                else:
                                    new_bqitem['locked'] = 0
                                new_rtitem['datatype'] = 2
                                itlist.append(new_rtitem)
                                num = num+1
                    datalist = itlist[start:end]
                    le = num
            else:
               datalist = {}
               le = 0
        response_data['data'] = serializer(datalist)
        response_data['draw'] = draw
        response_data['recordsTotal'] = le
        response_data['recordsFiltered'] = le
    except Exception as e:
        print e
        response_data['data'] = {}
        response_data['draw'] = 1
        response_data['recordsTotal'] = 0
        response_data['recordsFiltered'] = 0
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")

@login_required(login_url="/login/")
def chanzhibaobiao(request):
    templateName = 'TaskAndFlow/flowtemplate/business_shigongyusuan.html'
    pactmodals = Pact.objects.filter(type_id=1)[:1]
    if len(pactmodals)>0:
        for item in pactmodals:
            pactobj = item
    bqlist = BqItem.objects.filter().order_by('id')
    scpacts = PactFenbao.objects.filter(~Q(parent_code=''))
    companys = Company.objects.all()
    professionals = UserMajor.objects.all()
    pacttypes = PactType.objects.all()
    return render_to_response(templateName, RequestContext(request, locals()))

@login_required(login_url="/login/")
def chanzhioutandin(request):
    templateName = 'TaskAndFlow/flowtemplate/business_outandin.html'
    pactmodals = Pact.objects.filter(type_id=1)[:1]
    if len(pactmodals)>0:
        for item in pactmodals:
            pactobj = item
    bqlist = BqItem.objects.filter().order_by('id')
    companys = Company.objects.all()
    professionals = UserMajor.objects.all()
    pacttypes = PactType.objects.all()
    return render_to_response(templateName, RequestContext(request, locals()))

@login_required(login_url="/login/")
def ledgeraccount(request):
    templateName = 'TaskAndFlow/flowtemplate/business_account.html'
    version = time.time()
    if Report.objects.filter().count()>0:
        report_id = Report.objects.filter()[0].id
    else:
        report_id = 'null'
    bqlist = BqItem.objects.filter()
    if bqlist:
        bqitem = BqItem.objects.filter()[0]
    scpacts = PactFenbao.objects.filter(~Q(parent_code=''))
    resourcelist = RtItemResource.objects.filter()
    budgetlist = Budget.objects.filter()

    spacelist = PactSpace.objects.filter()
    pacts = Pact.objects.filter()
    majorlist = []
    comlist = []
    for item in pacts:
        majorlist.append(serializer(item.major))
        comlist.append(serializer(item.cooperation))

    majorarr = []
    comarr = []
    [comarr.append(x) for x in comlist if x not in comarr]#去重
    [majorarr.append(x) for x in majorlist if x not in majorarr]#去重

    return render_to_response(templateName, RequestContext(request, locals()))

''' 计算台账 '''
def manageraccount(request):
    response_data = {}
    response_data['issuc'] = True
    response_data['msg'] = ''
    atype = request.GET.get('type', '')
    space_id = request.GET.get('space_id', '')

    print space_id
    dataitem = {}
    tds = []
    listdata = []
    rnames = []
    if atype=='total':
        if space_id=='':
            tds = ReportRate.objects.filter(report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time']).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
        else:
            tds = ReportRate.objects.filter(report__pact__space_id=space_id,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report__pact__space_id=space_id,report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time'],report__pact__space_id=space_id).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
    elif atype=='zixing':
        if space_id=='':
            reports = Report.objects.filter(pact__is_self=1,locked=1).values_list('id')
            tds = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time'],report_id__in=reports).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
        else:
            reports = Report.objects.filter(pact__is_self=1,pact__space_id=space_id,locked=1).values_list('id')
            tds = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time'],report_id__in=reports).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
    elif atype=='fenbao':
        if space_id=='':
            reports = Report.objects.filter(pact__is_self=2,locked=1).values_list('id')
            tds = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time'],report_id__in=reports).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
        else:
            reports = Report.objects.filter(pact__is_self=2,pact__space_id=space_id,locked=1).values_list('id')
            tds = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = ReportRate.objects.filter(report_id__in=reports,report__locked=1).values('name','num').annotate(sum_money = Sum('money'))
            for item in rnames:
                item['tdlist'] =[]
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=item['num'],time = itemtd['time'],report_id__in=reports).aggregate(sum_money = Sum('money'))
                    item['tdlist'].append(tb)
                listdata.append(item)
    elif atype=='company':
        if space_id=='':
            tds = ReportRate.objects.filter(num=6,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = PactSpace.objects.filter()
            hejiitem = {}
            hejiitem['name'] = u'合计'
            hejiitem['sum_money'] = 0
            hejiitem['tdlist'] = []
            for itemA in rnames:
                item = serializer(itemA)
                item['tdlist'] =[]
                item['sum_money'] =0
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=6,time = itemtd['time'],report__pact__space_id=itemA.id).aggregate(sum_money = Sum('money'))
                    if tb['sum_money']!=None:
                        item['sum_money'] = item['sum_money']+tb['sum_money']
                    item['tdlist'].append(tb)
                hejiitem['sum_money'] = item['sum_money']+hejiitem['sum_money']
                listdata.append(item)

            for itemhj in tds:
                hejiitem['tdlist'].append({'sum_money':itemhj['sum_money']})
            listdata.append(hejiitem)
        else:
            tds = ReportRate.objects.filter(num=6,report__pact__space_id=space_id,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            rnames = PactSpace.objects.filter(id=space_id)
            hejiitem = {}
            hejiitem['name'] = u'合计'
            hejiitem['sum_money'] = 0
            hejiitem['tdlist'] = []
            for itemA in rnames:
                item = serializer(itemA)
                item['tdlist'] =[]
                item['sum_money'] =0
                for itemtd in tds:
                    tb = ReportRate.objects.filter(num=6,time = itemtd['time'],report__pact__space_id=itemA.id).aggregate(sum_money = Sum('money')).order_by('time')
                    if tb['sum_money']!=None:
                        item['sum_money'] = item['sum_money']+tb['sum_money']
                    item['tdlist'].append(tb)
                hejiitem['sum_money'] = item['sum_money']+hejiitem['sum_money']
                listdata.append(item)

            for itemhj in tds:
                hejiitem['tdlist'].append({'sum_money':itemhj['sum_money']})
            listdata.append(hejiitem)
    elif atype=='fenbaocompany':
        if space_id=='':
            tds = ReportRate.objects.filter(report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            companys = Pact.objects.filter(type=1).values('cooperation').annotate(Sum('id'))
            for item in companys:
                yusuanlist = []
                dataitem = {}
                dataitem['name'] = Company.objects.get(id=item['cooperation']).name
                dataitem['counts'] = Pact.objects.filter(cooperation_id=item['cooperation'],type=1).count()
                for index,companyitem in enumerate(Pact.objects.filter(cooperation_id=item['cooperation'],type=1)):
                    pitem = {'name':companyitem.name,'count':Report.objects.filter(pact_id=companyitem.id).count()}
                    pitem['sum_money'] = ReportRate.objects.filter(report__locked=1,num=6,report__pact_id=companyitem.id).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    pitem['rate'] = []
                    pitem['isshow'] = True if not index else False
                    # pitem['counts'] = count

                    for tditem in tds:
                        tdtime = tditem['time']
                        if Report.objects.filter(pact_id=companyitem.id,start_time=tdtime).count()>0:
                            reportitem = Report.objects.filter(pact_id=companyitem.id,start_time=tdtime)[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id).aggregate(sum_money = Sum('money'))
                            pitem['rate'].append(tb)
                        else:
                            pitem['rate'].append({'sum_money':None})

                    yusuanlist.append(pitem)
                dataitem['yusuanlist']  = yusuanlist
                listdata.append(dataitem)
        else:
            tds = ReportRate.objects.filter(report__pact__space_id=space_id,report__locked=1).values('time').annotate(sum_money = Sum('money')).order_by('time')
            companys = Pact.objects.filter(type=1,space_id=space_id).values('cooperation').annotate(Sum('id'))
            for item in companys:
                yusuanlist = []
                dataitem = {}
                dataitem['name'] = Company.objects.get(id=item['cooperation']).name
                dataitem['counts'] = Pact.objects.filter(cooperation_id=item['cooperation'],type=1,space_id=space_id).count()
                # count = Pact.objects.filter(cooperation_id=item['cooperation'],type=1,space_id=space_id).count()+1
                for index,companyitem in enumerate(Pact.objects.filter(cooperation_id=item['cooperation'],type=1)):
                    pitem = {'name':companyitem.name,'count':Report.objects.filter(pact_id=companyitem.id,pact__space_id=space_id).count()}
                    pitem['sum_money'] = ReportRate.objects.filter(report__locked=1,num=6,report__pact__space_id=space_id,report__pact_id=companyitem.id).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    pitem['rate'] = []
                    pitem['isshow'] = True if not index else False
                    # pitem['counts'] = count

                    for tditem in tds:
                        tdtime = tditem['time']
                        if Report.objects.filter(pact_id=companyitem.id,start_time=tdtime,pact__space_id=space_id).count()>0:
                            reportitem = Report.objects.filter(pact_id=companyitem.id,start_time=tdtime,pact__space_id=space_id)[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id,report__pact__space_id=space_id).aggregate(sum_money = Sum('money'))
                            pitem['rate'].append(tb)
                        else:
                            pitem['rate'].append({'sum_money':None})

                    yusuanlist.append(pitem)
                dataitem['yusuanlist']  = yusuanlist
                listdata.append(dataitem)
    elif atype=='zhuanye':
        tds = ReportRate.objects.filter(report__locked=1)
        companys = Pact.objects.filter(type=1)
        if space_id=='':
            tds = tds.values('time').annotate(sum_money = Sum('money')).order_by('time')
            companys = companys.values('major').annotate(Sum('id'))
            for item in companys:
                dataitem = {}
                dataitem['name'] = UserMajor.objects.get(id=item['major']).name
                dataitem['counts'] = Pact.objects.filter(major_id=item['major'],type=1).count()
                yusuanlist = []
                for index,companyitem in enumerate(Pact.objects.filter(major_id=item['major'],type=1)):
                    pitem = {'name':companyitem.name,'count':Report.objects.filter(pact_id=companyitem.id).count()}
                    pitem['sum_money'] = ReportRate.objects.filter(report__locked=1,num=6,report__pact__major_id=item['major'],report__pact_id=companyitem.id).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    print pitem['sum_money']
                    pitem['rate'] = []
                    pitem['isshow'] = True if not index else False
                    for tditem in tds:
                        tdtime = tditem['time']
                        if Report.objects.filter(pact_id=companyitem.id,start_time=tdtime).count()>0:
                            reportitem = Report.objects.filter(pact_id=companyitem.id,start_time=tdtime)[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id).aggregate(sum_money = Sum('money'))
                            pitem['rate'].append(tb)
                        else:
                            pitem['rate'].append({'sum_money':None})
                    yusuanlist.append(pitem)

                dataitem['yusuanlist']  = yusuanlist
                listdata.append(dataitem)
        else:
            tds = tds.filter(report__pact__space_id=space_id).values('time').annotate(sum_money = Sum('money')).order_by('time')
            companys = Pact.objects.filter(type=1,space_id=space_id).values('major').annotate(Sum('id'))
            for item in companys:
                dataitem = {}
                dataitem['name'] = UserMajor.objects.get(id=item['major']).name
                dataitem['counts'] = Pact.objects.filter(major_id=item['major'],type=1,space_id=space_id).count()
                yusuanlist = []
                for index,companyitem in enumerate(Pact.objects.filter(major_id=item['major'],type=1,space_id=space_id)):
                    pitem = {'name':companyitem.name,'count':Report.objects.filter(pact_id=companyitem.id,pact__space_id=space_id).count()}
                    pitem['sum_money'] = ReportRate.objects.filter(report__locked=1,num=6,report__pact__space_id=space_id,report__pact__major_id=item['major'],report__pact_id=companyitem.id).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    pitem['rate'] = []
                    pitem['isshow'] = True if not index else False
                    # pitem['counts'] = count
                    for tditem in tds:
                        tdtime = tditem['time']
                        if Report.objects.filter(pact_id=companyitem.id,start_time=tdtime,pact__space_id=space_id,pact__major_id=item['major']).count()>0:
                            reportitem = Report.objects.filter(pact_id=companyitem.id,start_time=tdtime,pact__space_id=space_id,pact__major_id=item['major'])[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id,report__pact__space_id=space_id,report__pact__major_id=item['major']).aggregate(sum_money = Sum('money'))
                            pitem['rate'].append(tb)
                        else:
                            pitem['rate'].append({'sum_money':None})
                    yusuanlist.append(pitem)

                dataitem['yusuanlist']  = yusuanlist
                listdata.append(dataitem)
    elif atype=='zidingyitz':
        start_time = request.GET.get('start_time', '')
        end_time = request.GET.get('end_time', '')
        company = request.GET.get('company', '')
        major = request.GET.get('major', '')
        zxorfb = request.GET.get('zxorfb', '')
        space = request.GET.get('space', '')

        pactlist = Pact.objects.filter(type=1)
        tds = ReportRate.objects.filter(report__locked=1).order_by('time')
        if company!='':
            tds = tds.filter(report__pact__cooperation_id__in=company.split(','))
            pactlist = pactlist.filter(cooperation_id__in=company.split(','))
        if space!='':
            tds = tds.filter(report__pact__space_id__in=space.split(','))
            pactlist = pactlist.filter(space_id__in=space.split(','))
        if major!='':
            tds = tds.filter(report__pact__major_id__in=major.split(','))
            pactlist = pactlist.filter(major_id__in=major.split(','))
        if zxorfb!='':
            tds = tds.filter(report__pact__is_self__in=zxorfb.split(','))
            pactlist = pactlist.filter(is_self__in=zxorfb.split(','))
        if start_time!='':
            tds = tds.filter(report__start_time__gte=datetime.datetime.strptime(start_time.split('-')[0].strip()+" 00:00:00",'%Y/%m/%d %H:%M:%S'))
        if end_time!='':
            tds = tds.filter(report__start_time__lte=datetime.datetime.strptime(end_time.split('-')[0].strip()+" 23:59:59",'%Y/%m/%d %H:%M:%S'))
        tds = tds.values('time').annotate(sum_money = Sum('money'))

        spacelist = pactlist.values('space').annotate(Sum('id'))

        for item in spacelist:
            print PactSpace.objects.get(id=item['space']).id
            dataitem = {}
            dataitem['name'] = PactSpace.objects.get(id=item['space']).name
            dataitem['counts'] = pactlist.filter(space_id=item['space']).count()
            dataitem['yusuanlist'] = []
            # dataitem['fbitem'] = {}
            yusuanlist = []
            if pactlist.filter(space_id=item['space'],is_self=1).count()>0:
                zxitem = {'name':u'自行','counts':pactlist.filter(space_id=item['space'],is_self=1).count()}
                zxitem['rate'] = []
                for i,y in enumerate(pactlist.filter(space_id=item['space'],is_self=1)):
                    yitem = {'name':y.name,'major':y.major.name,'company':y.cooperation.name}
                    yitem['sum_money'] = tds.filter(num=6).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    yitem['rs'] = tds.values('name','num').annotate(sum_money = Sum('money'))
                    yitem['isshow'] = True if not i else False
                    yitem['fenbaofirst'] = 0
                    yitem['rate'] = []
                    yitem['ratelist'] = {}
                    for tditem in tds:
                        tdtime = tditem['time']
                        rit = {}
                        if Report.objects.filter(pact_id=y.id,start_time=tdtime).count()>0:
                            reportitem = Report.objects.filter(pact_id=y.id,start_time=tdtime)[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id).aggregate(sum_money = Sum('money'))
                            rit['sum_money'] = tb['sum_money']
                            rit['ratelist'] = ReportRate.objects.filter(report_id=reportitem.id).values('name','num').annotate(sum_money = Sum('money'))
                        else:
                            rit['sum_money'] = None
                            rit['ratelist'] = None
                        yitem['rate'].append(rit)
                    zxitem['rate'].append(yitem)
                dataitem['yusuanlist'].append(zxitem)
            if pactlist.filter(space_id=item['space'],is_self=2).count()>0:
                fbitem = {'name':u'分包','counts':pactlist.filter(space_id=item['space'],is_self=2).count()}
                fbitem['rate'] = []
                for i,y in enumerate(pactlist.filter(space_id=item['space'],is_self=2)):
                    yitem = {'name':y.name,'major':y.major.name,'company':y.cooperation.name}
                    yitem['sum_money'] = tds.filter(num=6).aggregate(sum_money = Sum('money'))['sum_money']#合计
                    yitem['rs'] = tds.values('name','num').annotate(sum_money = Sum('money'))
                    yitem['fenbaofirst'] = 0
                    yitem['isshow'] = True if not i else False
                    if i == 0:
                        prec = 0
                        prec = pactlist.filter(space_id=item['space'],is_self=1).count()
                        if prec:
                            yitem['fenbaofirst'] = 1
                            if prec%2:
                                yitem['isshow'] = False
                    yitem['rate'] = []
                    for tditem in tds:
                        tdtime = tditem['time']
                        rit = {}
                        if Report.objects.filter(pact_id=y.id,start_time=tdtime).count()>0:
                            reportitem = Report.objects.filter(pact_id=y.id,start_time=tdtime)[0]
                            tb = ReportRate.objects.filter(num=6,report_id=reportitem.id).aggregate(sum_money = Sum('money'))
                            rit['sum_money'] = tb['sum_money']
                            rit['ratelist'] = ReportRate.objects.filter(report_id=reportitem.id).values('name','num').annotate(sum_money = Sum('money'))
                        else:
                            rit['sum_money'] = None
                            rit['ratelist'] = None
                        yitem['rate'].append(rit)
                    fbitem['rate'].append(yitem)
                dataitem['yusuanlist'].append(fbitem)
            listdata.append(dataitem)

    response_data['tds'] = serializer(tds)
    response_data['rnames'] = serializer(listdata)
    # print tds,listdata
    return HttpResponse(json.dumps(response_data), content_type="application/json")
def get_account_tree(request):
    try:
        id = request.GET.get('id', '')
        response_data = {}
        child_list = []

        if id == '#':
            response_data["id"] = "maint"
            response_data["text"] = "工程总台账"
            response_data["icon"] = "/img/buildings2.png"
            response_data["state"] = {'opened': True}
            treelist = PactSpace.objects.filter().order_by('order_no')

            for each in treelist:
                child_data = {}
                child_data["id"] = 'space-'+str(each.id)
                child_data["text"] = each.name
                child_data["icon"] = "/img/buildings2.png"
                child_data["state"] = {'opened': True}
                child_data_list = []
                children_datazx = {}
                children_datazx["id"] = 'zixing-'+str(each.id)
                children_datazx["text"] = '自行'
                children_datazx["icon"] = "/img/buildings2.png"
                # children_datazx["state"] = {'opened': True}
                children_datazx["children"] = True
                children_datafb = {}
                children_datafb["id"] = 'fenbao-'+str(each.id)
                children_datafb["text"] = '分包'
                children_datafb["icon"] = "/img/buildings2.png"
                # children_datafb["state"] = {'opened': True}
                children_datafb["children"] = True
                child_data_list.append(children_datafb)
                child_data_list.append(children_datazx)
                child_data["children"] = child_data_list
                child_list.append(child_data)

            response_data["children"] = child_list
        else:
            print id
            idarr = id.split('-')
            print idarr
            child_list = []
            if idarr[0] == 'fenbao':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=2):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/hetong.png"
                    pactitem_data["state"] = {'opened': True}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)
            elif idarr[0] == 'zixing':
                pactitemlist = []
                for pactitem in Pact.objects.filter(space_id=idarr[1],is_self=1):
                    pactitem_data = {}
                    pactitem_data["id"] = 'zixingyg-'+idarr[1]+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/hetong.png"
                    pactitem_data["state"] = {'opened': True}
                    pactitem_data['children'] = True
                    child_list.append(pactitem_data)


            response_data = child_list

        return HttpResponse(json.dumps(response_data), content_type="application/json")
    except:
        traceback.print_exc()


def tasktreemanager(pactid):
    reportlist = Report.objects.filter(pact_id=pactid).order_by('-start_time')#分包合同
    child_list = []
    if len(reportlist)>0:
        for each in reportlist:
            print each.name
            child_data = {}
            child_physical_data = {}#实物量任务单
            child_physical_data["id"] = 'allyesphysical-'+str(each.id)
            child_physical_data["text"] = (each.name).replace(u'产值',u'实物量任务单')
            child_physical_data["icon"] = "/img/hetong.png"

            child_nophysical_data = {}#非实物量任务单
            child_nophysical_data["id"] = 'allnophysical-'+str(each.id)
            child_nophysical_data["text"] = (each.name).replace(u'产值',u'非实物量任务单')
            child_nophysical_data["icon"] = "/img/hetong.png"

            costlist = ReportCostItem.objects.filter(report_id=each.id)
            company_physical_list = []
            company_nophysical_list = []
            res = []
            #获取其让单位的分建成本和劳务成本
            physicallist = []
            nophysicallist = []
            if TaskOrderPhysical.objects.filter(report_id=each.id).count()>0:
                physicallist = TaskOrderPhysical.objects.filter(report_id=each.id).order_by('company_id')
                cal_list = []
                curr_company_id = 0
                pre_company_id = 0
                company_data = {}#公司
                conpany_children_list = []
                index = 1
                for item in physicallist:
                    curr_company_id = item.company.id
                    if pre_company_id==0:
                        pre_company_id = curr_company_id
                        company_data["id"] = 'companyyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        company_data["text"] = item.company.name
                        company_data["icon"] = "/img/hetong.png"
                        company_data["state"] = {'opened': True}

                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'taskyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.sc.name
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)
                        if index == len(physicallist):
                            company_data["children"] = conpany_children_list
                            cal_list.append(company_data)
                    elif pre_company_id==curr_company_id:
                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'taskyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.sc.name
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)
                        if index == len(physicallist):
                            company_data["children"] = conpany_children_list
                            company_data["id"] = 'companyyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                            company_data["text"] = item.company.name
                            company_data["icon"] = "/img/hetong.png"
                            company_data["state"] = {'opened': True}
                            cal_list.append(company_data)
                    else:
                        company_data["children"] = conpany_children_list
                        cal_list.append(company_data)
                        pre_company_id = curr_company_id
                        company_data = {}#公司
                        company_data["id"] = 'companyyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        company_data["text"] = item.company.name
                        company_data["icon"] = "/img/hetong.png"
                        company_data["state"] = {'opened': True}

                        conpany_children_list = []
                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'taskyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.sc.name
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)

                        if index == len(physicallist):
                            company_data["children"] = conpany_children_list
                            company_data["id"] = 'companyyesphysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                            company_data["text"] = item.company.name
                            company_data["icon"] = "/img/hetong.png"
                            company_data["state"] = {'opened': True}
                            cal_list.append(company_data)
                    index = index+1
                company_physical_list = cal_list

            if TaskOrderNophysical.objects.filter(report_id=each.id).count()>0:

                nophysicallist = TaskOrderNophysical.objects.filter(report_id=each.id)
                cal_list = []
                curr_company_id = 0
                pre_company_id = 0
                company_data = {}#公司
                conpany_children_list = []
                index = 1
                for item in nophysicallist:
                    # print u'有数据'+str(item.id)
                    curr_company_id = item.company.id
                    if pre_company_id==0:
                        pre_company_id = curr_company_id
                        company_data["id"] = 'companynophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        company_data["text"] = item.company.name
                        company_data["icon"] = "/img/hetong.png"
                        company_data["state"] = {'opened': True}

                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'tasknophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.worktype
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)

                        if index == len(nophysicallist):
                            company_data["children"] = conpany_children_list
                            cal_list.append(company_data)
                    elif pre_company_id==curr_company_id:
                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'tasknophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.worktype
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)

                        if index == len(nophysicallist):
                            company_data["children"] = conpany_children_list
                            company_data["id"] = 'companynophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                            company_data["text"] = item.company.name
                            company_data["icon"] = "/img/hetong.png"
                            company_data["state"] = {'opened': True}
                            cal_list.append(company_data)
                    else:
                        company_data["children"] = conpany_children_list
                        cal_list.append(company_data)
                        pre_company_id = curr_company_id
                        company_data = {}#公司
                        company_data["id"] = 'companynophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        company_data["text"] = item.company.name
                        company_data["icon"] = "/img/hetong.png"
                        company_data["state"] = {'opened': True}


                        conpany_children_list = []
                        task_item_data = {}#实物量任务单
                        task_item_data["id"] = 'tasknophysical-'+str(each.id)+'-'+str(item.company.id)+'-'+str(item.id)
                        task_item_data["text"] = item.worktype
                        task_item_data["icon"] = "/img/hetong.png"
                        task_item_data["state"] = {'opened': True}
                        conpany_children_list.append(task_item_data)

                        if index == len(nophysicallist):
                            company_data["children"] = conpany_children_list
                            company_data["id"] = 'companynophysical-'+str(each.id)+'-'+str(item.company.id)
                            company_data["text"] = item.company.name
                            company_data["icon"] = "/img/hetong.png"
                            company_data["state"] = {'opened': True}
                            cal_list.append(company_data)
                    index = index+1
                company_nophysical_list = cal_list

            child_physical_data['children'] = company_physical_list
            child_nophysical_data['children'] = company_nophysical_list
            child_list.append(child_physical_data)
            child_list.append(child_nophysical_data)
    else:
        child_list = []

    return child_list

def getspaceaccountlist(space_id):
    listdata = []
    return listdata

from django.http import HttpResponse
from xlwt import *
def excel_export(request):
    """
    导出excel表格
    """
    list_obj = Comment.objects.all().order_by("-time")
    if list_obj:
        # 创建工作薄
        ws = Workbook(encoding='utf-8')
        w = ws.add_sheet(u"数据报表第一页")
        w.write(0, 0, "id")
        w.write(0, 1, u"用户名")
        w.write(0, 2, u"发布时间")
        w.write(0, 3, u"内容")
        w.write(0, 4, u"来源")
        # 写入数据
        excel_row = 1
        for obj in list_obj:
            data_id = obj.id
            data_user = obj.username
            data_time = obj.time.strftime("%Y-%m-%d")[:10]
            data_content = obj.content
            dada_source = obj.source
            w.write(excel_row, 0, data_id)
            w.write(excel_row, 1, data_user)
            w.write(excel_row, 2, data_time)
            w.write(excel_row, 3, data_content)
            w.write(excel_row, 4, dada_source)
            excel_row += 1
        # 检测文件是够存在
        # 方框中代码是保存本地文件使用，如不需要请删除该代码
        ###########################
        exist_file = os.path.exists("test.xls")
        if exist_file:
            os.remove(r"test.xls")
        ws.save("test.xls")
        ############################
        sio = StringIO.StringIO()
        ws.save(sio)
        sio.seek(0)
        response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=test.xls'
        response.write(sio.getvalue())
        return response


####其他预算
@login_required(login_url="/login/")
def otheritem_yusuan(request):
    templateName = 'TaskAndFlow/flowtemplate/business_otheritem.html'
    version = time.time()
    try:

        print 1111
    except Exception as e:
        traceback.print_exc()
    return render_to_response(templateName, RequestContext(request, locals()))

def get_business_otheritem_tree(request):
    try:
        id = request.GET.get('id', '')
        response_data = {}
        child_list = []
        if id == '#':
            response_data["id"] = "maint"
            response_data["text"] = "智慧建造平台"
            response_data["icon"] = "/img/buildings2.png"
            response_data["state"] = {'opened': True}
            treelist = PactSpace.objects.filter().order_by('order_no')
            for each in treelist:
                child_data = {}
                child_data["id"] = 'space-'+str(each.id)
                child_data["text"] = each.name
                child_data["icon"] = "/img/buildings2.png"
                child_data["state"] = {'opened': True}
                child_data_list = []
                children_datazx = {}
                children_datazx["id"] = 'zixing-'+str(each.id)
                children_datazx["text"] = '自行'
                children_datazx["icon"] = "/img/buildings2.png"
                children_datazx["children"] = True
                children_datafb = {}
                children_datafb["id"] = 'fenbao-'+str(each.id)
                children_datafb["text"] = '分包'
                children_datafb["icon"] = "/img/buildings2.png"
                children_datafb["children"] = True
                child_data_list.append(children_datafb)
                child_data_list.append(children_datazx)
                child_data["children"] = child_data_list
                child_list.append(child_data)

            response_data["children"] = child_list
        else:
            idarr = id.split('-')
            child_list = []
            if idarr[0] == 'fenbao' or idarr[0] == 'zixing':
                pactlist = []
                node_id = ''
                if idarr[0] == 'fenbao':
                    pactlist = Pact.objects.filter(space_id=idarr[1],is_self=2)
                    node_id = 'zixingyg-'+idarr[1]
                elif idarr[0] == 'zixing':
                    pactlist = Pact.objects.filter(space_id=idarr[1],is_self=1)
                    node_id = 'zixingyg-'+idarr[1]

                for pactitem in pactlist:
                    pactitem_data = {}
                    pactitem_data["id"] = node_id+'-'+str(pactitem.id)
                    pactitem_data["text"] = pactitem.name
                    pactitem_data["icon"] = "/img/hetong.png"
                    pactitem_data["state"] = {'opened': False}
                    pactitem_data['children'] = []

                    fenbufenx = {}
                    fenbufenx["id"] = 'fenbufenxiang-'+str(pactitem.id)
                    fenbufenx["text"] = u'分部分项清单'
                    fenbufenx["icon"] = "/img/hetong.png"
                    fenbufenx["state"] = {'opened': True}
                    feilvbiao = {}
                    feilvbiao["id"] = 'feilvbiao-'+str(pactitem.id)
                    feilvbiao["text"] = u'费率表'
                    feilvbiao["icon"] = "/img/hetong.png"
                    feilvbiao["state"] = {'opened': True}

                    pactitem_data['children'].append(fenbufenx)
                    pactitem_data['children'].append(feilvbiao)
                    child_list.append(pactitem_data)

            response_data = child_list

        return HttpResponse(json.dumps(response_data), content_type="application/json")
    except:
        traceback.print_exc()


@login_required(login_url="/login/")
def costManager_vue(request):
    return render_to_response('costManager/costManager.html', RequestContext(request,locals()))


'''台账数据接口'''
def loadAccountList(request):
    response_data = {}
    try:
        response_data['issuc'] = True
        operate = request.GET.get('type')
        pact_id = request.GET.get('pact_id')
        print pact_id
        bqlist = BqItem.objects.filter(pact_id=pact_id)
        reports = Report.objects.filter(pact_id=pact_id,locked=1).order_by('start_time')
        reportlist = []
        for report in reports:
            print report.name
            time = report.start_time.strftime('%Y-%m')
            reportlist.append(time)
        tds = ReportRate.objects.filter(report__pact_id=pact_id,report__locked=1).values('time').annotate(sum_money = Sum('money'))

        response_data['reportlist'] = reportlist
        datalist = []
        for bqitem in BqItem.objects.filter(pact_id=pact_id):
            bqitemdata = {}
            bqitemdata['id'] = bqitem.id
            bqitemdata['code'] = bqitem.BQItem_Code
            bqitemdata['name'] = bqitem.BQItemName
            bqitemdata['unit'] = bqitem.BQItemUnit
            bqitemdata['quantities'] = bqitem.buildBqs
            bqitemdata['price'] = bqitem.allunitrate
            bqitemdata['complete_quantities'] = 0 #累计完成
            bqitemdata['remainder_money'] = 0 #剩余金额
            bqitemdata['remainder_quantities'] = 0 #剩余量
            bqitemdata['reportlist']= []
            for report in reports:
                try:
                    rbitem = ReportBqitem.objects.get(report=report,bq=bqitem)
                    bqitemdata['complete_quantities'] = bqitemdata['complete_quantities']+float(rbitem.quantity)
                    bqitemdata['remainder_money'] = bqitemdata['remainder_money']+float(rbitem.quantity)*float(rbitem.unitprice)
                    bqitemdata['reportlist'].append([rbitem.quantity,float(rbitem.unitprice)*float(rbitem.quantity)])
                except:
                    bqitemdata['reportlist'].append([0,0])
            bqitemdata['remainder_money'] = bqitem.buildBqs*float(bqitem.allunitrate)-bqitemdata['remainder_money']
            bqitemdata['remainder_quantities'] = bqitem.buildBqs-bqitemdata['complete_quantities']
            datalist.append(bqitemdata)
            for rtitem in RtItem.objects.filter(BQ_Item_id=bqitem.id):
                rtitemdata = {}
                rtitemdata['id'] = rtitem.id
                rtitemdata['code'] = rtitem.RQItem_Code
                rtitemdata['name'] = rtitem.RT_Item_Name
                rtitemdata['unit'] = rtitem.unit
                rtitemdata['quantities'] = rtitem.quantities*bqitem.buildBqs/bqitem.designBqs
                rtitemdata['price'] = rtitem.unitprice
                rtitemdata['complete_quantities'] = 0 #累计完成
                rtitemdata['remainder_money'] = 0 #剩余金额
                rtitemdata['reportlist'] = []
                for report in reports:
                    try:
                        rbitem = ReportRtitem.objects.get(reportbq__report=report,rtitem=rtitem)
                        rtitemdata['complete_quantities'] = rtitemdata['complete_quantities']+float(rbitem.quantity)
                        rtitemdata['remainder_money'] = rtitemdata['remainder_money']+float(rbitem.unitprice)*float(rbitem.quantity)
                        rtitemdata['reportlist'].append([rbitem.quantity,float(rbitem.unitprice)*float(rbitem.quantity)])
                    except:
                        rtitemdata['reportlist'].append([0,0])
                rtitemdata['remainder_money'] = rtitemdata['quantities']*float(rtitem.unitprice)-rtitemdata['remainder_money']
                rtitemdata['remainder_quantities'] = rtitemdata['quantities']-rtitemdata['complete_quantities'] #剩余量
                datalist.append(rtitemdata)
        response_data['datalist'] = serializer(datalist)
    except Exception as e:
        response_data['issuc'] = False
        traceback.print_exc()
    return HttpResponse(json.dumps(response_data,encoding="UTF-8",ensure_ascii=False), content_type="application/json")
