# -*- coding: utf-8 -*-
# Create your views here.
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.shortcuts import render_to_response
from models import *
import os, traceback, sys
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
import json,thread, base64,codecs
from TaskAndFlow.utility import *
from TaskAndFlow.utility_filemanager import *
from TaskAndFlow.utility_taskmanager import *
from django.core import serializers
from django.db.models import Q
from Scc4PM.settings import UPLOAD_DIR
from openpyxl import Workbook
from openpyxl import load_workbook
from UserAndPrj.models import *
from UserAndPrj.utility import *

def handle_uploaded_file(f, name="",filepath=""):
    file_name = ""

    try:
        path=settings.UPLOAD_DIR
        print path
        if not os.path.exists(path):
            os.makedirs(path)
        if name !="":
            oldname = name
        else:
            oldname = os.path.splitext(f.name)[0]
        ext = os.path.splitext(f.name)[1]

        #定义文件名，年月日时分秒随机数
        fn = time.strftime('%Y%m%d%H%M%S')
        fn =  '_'+fn
        #重写合成文件名
        newname =oldname+fn + ext
        file_name = path + newname
        #print file_name
        destination = open(file_name, 'wb+')
        for chunk in f.chunks():
            destination.write(chunk)
        destination.close()
    except Exception, e:
        print e

    return oldname,newname


@csrf_exempt
def uploadfile_conc(request):
    response_data = {}
    response_data["issuc"]="false"
    normal = request.POST.get('doctype','normal')
    print normal
    try:
        fileinfo = request.FILES.get("Filedata",None)
        # if not fileinfo:
        #     fileinfo = request.FILES.get('file',None)
        if fileinfo:
            name,filename=handle_uploaded_file(fileinfo)
            doc = Document()
            doc.name = filename
            doc.shortname = getfilerealname(fileinfo.name)
            doc.filepath = u"upload/"
            doc.creator=request.user
            doc.filesize = fileinfo._size
            doc.filetype = getfiletype(fileinfo)
            doc.doctype=normal
            doc.save()

            response_data["issuc"]="true"
            response_data["docId"]=doc.id
            response_data["docpath"]='/'+str(doc.filepath)+doc.name
            response_data["docname"]=doc.name

    except Exception, e:
        response_data["issuc"]="false"
        print e

    return HttpResponse(json.dumps(response_data), content_type="application/json")

@csrf_exempt
def del_uploadfile(request):
    try:
        fileid= request.POST.get('fileid')

        doc = Document.objects.get(id=int(fileid))
        path=settings.UPLOAD_DIR+'/'.join(str(doc.filepath).split('/')[1:])+doc.name
        print path
        os.remove(path)
        doc.delete()
        return HttpResponse(json.dumps({"issuc":"true"}), content_type="application/json")
    except:
        traceback.print_exc()

@csrf_exempt
def uploadfile_conc2(request):
    response_data = {}
    response_data["issuc"]="false"
    try:
        files = request.FILES.getlist('fileList')
        if not files:
            files = request.FILES.getlist('files[]')
        if not files:
            files = request.FILES.getlist('file')

        filedir = request.POST.get("filedir",'')
        if filedir:
            filedir='/'.join(filedir.split('/')[:-1])
            print filedir

        for fileinfo in files:
            name,filename=handle_uploaded_file(fileinfo)
            doc = Document()
            doc.name = filename
            doc.shortname = getfilerealname(fileinfo.name)
            doc.filepath = u"upload/"
            doc.creator=request.user
            doc.filesize = fileinfo._size
            doc.filetype = getfiletype(fileinfo)
            doc.doctype='normal'
            doc.save()

            response_data["issuc"]="true"
            response_data["docId"]=doc.id
            response_data["url"]="/upload/"+filename
            response_data["filedir"]=filedir
            print json.dumps(response_data)
    except:
        traceback.print_exc()
        response_data["issuc"]="false"


    return HttpResponse(json.dumps(response_data), content_type="application/json")


@csrf_exempt
def uploadfile_blob(request):
    response_data = {}
    response_data["issuc"]="false"
    try:
        imagedata  = request.POST.get('imgdata')
        imagedata  = base64.b64decode(imagedata) 
       
        path=settings.UPLOAD_DIR

        weekly_begindate,weekly_enddate = getWeeklyDate()
        timerange = weekly_begindate.strftime("%Y.%m.%d")+"-"+weekly_enddate.strftime("%Y.%m.%d")
        filename = u"工程周报("+timerange+u").png"
        file_path_name = path + filename

        destination = open(file_path_name, 'wb+')
        destination.write(imagedata)
        destination.close()

        if not Document.objects.filter(name=filename):
            doc = Document()
            doc.name = filename
            doc.shortname = filename
            doc.filepath = u"upload/"
            doc.creator=request.user
            doc.filesize = len(imagedata)
            doc.filetype = "image/png"
            doc.doctype='normal'
            doc.save()

            weeklydir = getWeeklyDir()        
            if weeklydir:
                doc.docdirectory.add(weeklydir)
                movefiletoDir(doc,weeklydir)

            weekly = getCurWeekly(request.user)
            weekly.file=doc
            weekly.save()

        response_data["issuc"]="true"

    except:
        traceback.print_exc()
        response_data["issuc"]="false"


    return HttpResponse(json.dumps(response_data), content_type="application/json")




@csrf_exempt
def uploadfile_ocf(request):
    response_data = {}
    response_data["issuc"]="false"
    try:
        files = request.FILES.getlist('fileList')
        if not files:
            files = request.FILES.getlist('files[]')
        if not files:
            files = request.FILES.getlist('file')

        fileid = request.POST.get("fileid",None)
        if fileid:
            dwgfile = Document.objects.get(id=fileid)

        for fileinfo in files:
            name,filename=handle_uploaded_file(fileinfo)
            dwgfile.previewfile = filename
            dwgfile.save()

        response_data["issuc"]="true"
    except:
        traceback.print_exc()
        response_data["issuc"]="false"


    return HttpResponse(json.dumps(response_data), content_type="application/json")

def profile_upload_com(file):
    '''文件上传函数'''
    if file:
        path=os.path.join(settings.UPLOAD_DIR,'tmp/')
        file_name=file.name
        #print file_name
        #fname = os.path.join(settings.MEDIA_ROOT,filename)
        path_file=os.path.join(path,file_name)
        fp = open(path_file, 'wb+')
        for content in file.chunks():
            fp.write(content)
        fp.close()

        return (True,file_name) #change
    return (False,file_name)   #change

#用户管理-添加用户-删除附件

@csrf_exempt
def profile_delte(request):
    del_file=request.POST.get("delete_file",'')
    if del_file:
        path=os.path.join(settings.UPLOAD_DIR,'tmp/')
        path_file=os.path.join(path,del_file)
        os.remove(path_file)

@csrf_exempt
def profile_delte_doc(request):
    ret={}
    del_file=request.POST.get("name",'')
    if del_file:
        path=settings.UPLOAD_DIR
        path_file=os.path.join(path,os.path.basename(del_file))

        if os.path.isfile(path_file):
            os.remove(path_file)

        try:
            Document.objects.get(content=del_file).delete()
            ret["status"] = "Succeed"

        except Document.DoesNotExist:
            ret["status"] = "Fail"
            ret["msg"] = "无记录！"

    else:
        ret["status"] = "Fail"
        ret["msg"] = "不能获取目标文件！"

    return HttpResponse(json.dumps(ret,ensure_ascii = False))

@login_required(login_url="/login/") 
@csrf_exempt
def importusersfile(request):
    response_data = {}
    response_data["issuc"]=False
    response_data["faillist"] = []
    response_data["total"] = 0
    response_data["suctotal"] = 0
    try:
        fileinfo = request.FILES.get("Filedata",None)

        if fileinfo:
            name,filename=handle_uploaded_file(fileinfo)
            #filename = u"用户导入模板.xlsx"
            path=settings.UPLOAD_DIR
            filepath = path + filename
            wb = load_workbook(filepath)
            sheet = wb.get_sheet_by_name(wb.sheetnames[0])
            #数据有效行从第2行开始        
            start_row = 2
            rows_len = sheet.max_row         #行数
            columns_len = sheet.max_column    #列数
            if columns_len>8:
                raise Exception("模板格式不正确，列数太多!")

            if rows_len<2:
                raise Exception("没有用户需要导入，请填写要导入的用户!")
            
            faillist = []
            suctotal = 0
            for row in range(start_row, rows_len+1):
                tmpinfo = {}
                for column in range(1,columns_len+1):
                    if 1==column:
                        tmpinfo["username"]=sheet.cell(row=row,column=column).value
                    elif 2==column:
                        tmpinfo["truename"]=sheet.cell(row=row,column=column).value
                    elif 3==column:
                        tmpinfo["password"]=str(sheet.cell(row=row,column=column).value)
                    elif 4==column:
                        tmpinfo["company"]=sheet.cell(row=row,column=column).value
                    elif 5==column:
                        tmpinfo["major"]=sheet.cell(row=row,column=column).value
                    elif 6==column:
                        tmpinfo["division"]=sheet.cell(row=row,column=column).value
                    elif 7==column:
                        tmpinfo["roles"]=sheet.cell(row=row,column=column).value
                    elif 8==column:
                        tmpinfo["contract"]=str(sheet.cell(row=row,column=column).value)


                if User.objects.filter(name=tmpinfo["username"]):
                    tmpinfo["error"] = u"用户名已存在"
                    faillist.append(tmpinfo)
                elif len(tmpinfo["password"])>12 or len(tmpinfo["password"])<6:
                    tmpinfo["error"] = u"密码长度不对（6-12）"
                    faillist.append(tmpinfo)
                elif tmpinfo["company"] and not Company.objects.filter(name=tmpinfo["company"]):
                    tmpinfo["error"] = u"公司不存在，请先添加对应公司"
                    faillist.append(tmpinfo)
                elif tmpinfo["major"] and not UserMajor.objects.filter(name=tmpinfo["major"]):
                    tmpinfo["error"] = u"专业不存在，请先添加对应专业"
                    faillist.append(tmpinfo)
                elif tmpinfo["division"] and not Division.objects.filter(name=tmpinfo["division"]):
                    tmpinfo["error"] = u"参加方不存在，请先添加对应参加方"
                    faillist.append(tmpinfo)
                elif not tmpinfo["roles"]:
                    tmpinfo["error"] = u"用户角色不能为空"
                    faillist.append(tmpinfo)
                else:
                    try:
                        company = None
                        if tmpinfo["company"]:
                            company = Company.objects.get(name=tmpinfo["company"]).id

                        major = None
                        if tmpinfo["major"]:
                            major = UserMajor.objects.get(name=tmpinfo["major"]).id
              
                        division = None
                        if tmpinfo["division"]:
                            division = Division.objects.get(name=tmpinfo["division"]).id

                        rolelist = tmpinfo["roles"].split("，")
                        roleidlist = []
                        if len(rolelist)==1:
                            rolelist = tmpinfo["roles"].split(",")
                        print rolelist
                        try:
                            for each in rolelist:
                                roleidlist.append(Role.objects.get(name=each).id)
                        except Exception as e:
                            raise Exception("角色不存在，请先添加对应角色!")

                        CreateOneUser(tmpinfo["username"],tmpinfo["contract"],tmpinfo["truename"],False,
                                        company,major, division,tmpinfo["password"],roleidlist)
                        suctotal +=1
                        print tmpinfo
                    except Exception, e:
                        tmpinfo["error"] = '%s' % e
                        faillist.append(tmpinfo)

            response_data["faillist"] = faillist
            response_data["total"] = rows_len-1
            response_data["suctotal"] = suctotal
            response_data["issuc"]=True
    except Exception, e:
        traceback.print_exc()
        response_data['error'] = '%s' % e

    return HttpResponse(json.dumps(response_data), content_type="application/json")



